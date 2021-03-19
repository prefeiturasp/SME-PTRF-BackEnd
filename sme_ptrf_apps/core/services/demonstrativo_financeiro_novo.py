import logging
import os

import re
from copy import copy
from tempfile import NamedTemporaryFile

from datetime import date

from django.db.models import Sum
from django.contrib.staticfiles.storage import staticfiles_storage
from django.core.files import File
from openpyxl import load_workbook, styles
from sme_ptrf_apps.core.services.xlsx_copy_row import copy_row, insert_row

from sme_ptrf_apps.core.choices import MembroEnum

from sme_ptrf_apps.core.models import (FechamentoPeriodo, MembroAssociacao, ObservacaoConciliacao,
                                       DemonstrativoFinanceiro)
from sme_ptrf_apps.despesas.models import RateioDespesa
from sme_ptrf_apps.receitas.models import Receita
from sme_ptrf_apps.receitas.tipos_aplicacao_recurso_receitas import (APLICACAO_CAPITAL, APLICACAO_CUSTEIO,
                                                                     APLICACAO_LIVRE)

global OFFSET

LOGGER = logging.getLogger(__name__)

COL_CABECALHO = 10
LINHA_PERIODO_CABECALHO = 4
LINHA_CONTA_CABECALHO = 5
LAST_LINE = 64
OFFSET = 0

# Coluna 2 da planilha
SALDO_ANTERIOR = 2
CREDITO = 3
DESPESA_REALIZADA = 4
DESPESA_NAO_REALIZADA = 5
SALDO_REPROGRAMADO_PROXIMO = 6
TOTAL_REPROGRAMADO_PROXIMO = 7
DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS = 8
SALDO_BANCARIO = 9
TOTAL_SALDO_BANCARIO = 10

# Colunas dos blocos 4 e 5 da planilha
ITEM = 0
RAZAO_SOCIAL = 1
CNPJ_CPF = 2
TIPO_DOCUMENTO = 3
NUMERO_DOCUMENTO = 4
DATA = 5
ACAO_DOCUMENTO = 6
ESPECIFICACAO_MATERIAL = 7
TIPO_DESPESA = 8
TIPO_TRANSACAO = 9
DATA_2 = 10
VALOR = 11


def gerar_arquivo_demonstrativo_financeiro_novo(acoes, periodo, conta_associacao, prestacao, usuario=""):
    filename = 'demonstrativo_financeiro_%s.xlsx'

    xlsx = gerar(usuario, acoes, periodo, conta_associacao)

    with NamedTemporaryFile() as tmp:
        xlsx.save(tmp.name)

        demonstrativo_financeiro, _ = DemonstrativoFinanceiro.objects.update_or_create(
            conta_associacao=conta_associacao,
            prestacao_conta=prestacao
        )

        demonstrativo_financeiro.arquivo.save(name=filename % demonstrativo_financeiro.pk, content=File(tmp))


def gerar(usuario, acoes, periodo, conta_associacao, previa=False):
    try:
        LOGGER.info("GERANDO DEMONSTRATIVO...")
        rateios_conferidos = RateioDespesa.rateios_da_conta_associacao_no_periodo(
            conta_associacao=conta_associacao, periodo=periodo, conferido=True)

        rateios_nao_conferidos = RateioDespesa.rateios_da_conta_associacao_no_periodo(
            conta_associacao=conta_associacao, periodo=periodo, conferido=False)

        rateios_nao_conferidos_em_periodos_anteriores = RateioDespesa.rateios_da_conta_associacao_em_periodos_anteriores(
            conta_associacao=conta_associacao, periodo=periodo, conferido=False)

        receitas_demonstradas = Receita.receitas_da_conta_associacao_no_periodo(
            conta_associacao=conta_associacao, periodo=periodo, conferido=True)

        fechamento_periodo = FechamentoPeriodo.objects.filter(
            conta_associacao=conta_associacao, periodo__uuid=periodo.uuid).first()

        path = os.path.join(os.path.basename(staticfiles_storage.location), 'cargas')
        nome_arquivo = os.path.join(path, 'modelo_demonstrativo_financeiro_novo_v2.xlsx')
        workbook = load_workbook(nome_arquivo)
        worksheet = workbook.active

        cabecalho(worksheet, periodo, conta_associacao, previa)

        bloco1_identificacao_apm(worksheet, acoes)
        bloco2_identificacao_conta(worksheet, conta_associacao)
        bloco3_resumo_por_acao(worksheet, acoes, conta_associacao, periodo, fechamento_periodo)
        bloco4_creditos_demonstrados(worksheet, receitas_demonstradas)
        bloco5_despesas_demonstradas(worksheet, rateios_conferidos)
        bloco6_despesas_demonstradas(worksheet, rateios_nao_conferidos)
        bloco7_despesas_demonstradas(worksheet, rateios_nao_conferidos_em_periodos_anteriores)
        #observacoes(worksheet, acao_associacao, periodo, conta_associacao)
        data_geracao_documento(worksheet, usuario, previa)
    # except Exception as e:
    #    LOGGER.error("ERRO no Demonstrativo: %s", str(e))
    finally:
        LOGGER.info("DEMONSTRATIVO GERADO")

    return workbook


def cabecalho(worksheet, periodo, conta_associacao, previa):
    global OFFSET
    OFFSET = 0

    rows = list(worksheet.rows)
    texto = "Demonstrativo Financeiro - PRÉVIA" if previa else "Demonstrativo Financeiro - FINAL"
    rows[2][7].value = texto
    rows[LINHA_PERIODO_CABECALHO][COL_CABECALHO].value = str(periodo)
    rows[LINHA_CONTA_CABECALHO][COL_CABECALHO].value = conta_associacao.tipo_conta.nome


def bloco1_identificacao_apm(worksheet, acoes):
    """BLOCO 1 - IDENTIFICAÇÃO DA APM/APMSUAC DA UNIDADE EDUCACIONAL"""
    acao_associacao = acoes[0]
    associacao = acao_associacao.associacao
    rows = list(worksheet.rows)
    rows[10][0].value = associacao.nome
    rows[10][6].value = associacao.cnpj
    rows[10][8].value = associacao.unidade.codigo_eol or ""
    rows[10][9].value = associacao.unidade.dre.nome if associacao.unidade.dre else ""

    presidente_diretoria_executiva = MembroAssociacao.objects.filter(associacao=associacao,
                                                                     cargo_associacao=MembroEnum.PRESIDENTE_DIRETORIA_EXECUTIVA.name).first()

    presidente_conselho_fiscal = MembroAssociacao.objects.filter(associacao=associacao,
                                                                 cargo_associacao=MembroEnum.PRESIDENTE_CONSELHO_FISCAL.name).first()

    rows[LAST_LINE - 3][0].value = presidente_diretoria_executiva.nome if presidente_diretoria_executiva else ''
    rows[LAST_LINE - 3][6].value = presidente_conselho_fiscal.nome if presidente_conselho_fiscal else ''


def bloco2_identificacao_conta(worksheet, conta_associacao):
    rows = list(worksheet.rows)

    banco = conta_associacao.banco_nome
    agencia = conta_associacao.agencia
    conta = conta_associacao.numero_conta
    data_extrato = date.today().strftime("%d/%m/%Y")
    saldo_extrato = 0

    row = rows[15]
    row[0].value = banco
    row[3].value = agencia
    row[5].value = conta
    row[7].value = data_extrato
    row[9].value = saldo_extrato


def bloco3_resumo_por_acao(worksheet, acoes, conta_associacao, periodo, fechamento_periodo):
    global OFFSET

    linha_inicial = 21 + OFFSET
    linha_atual = linha_inicial
    offset_local = 0

    total_valores = 0
    total_conciliacao = 0

    destinacoes = ['C', 'K', 'CK']

    totalizador = {
        SALDO_ANTERIOR: {'C': 0, 'K': 0, 'CK': 0},
        CREDITO: {'C': 0, 'K': 0, 'CK': 0},
        DESPESA_REALIZADA: {'C': 0, 'K': 0, 'CK': 0},
        DESPESA_NAO_REALIZADA: {'C': 0, 'K': 0, 'CK': 0},
        SALDO_REPROGRAMADO_PROXIMO: {'C': 0, 'K': 0, 'CK': 0},
        DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS: {'C': 0, 'K': 0, 'CK': 0},
        SALDO_BANCARIO: {'C': 0, 'K': 0, 'CK': 0},
        TOTAL_SALDO_BANCARIO: {'C': 0, 'K': 0, 'CK': 0},
    }

    logging.info(f'Linha inicial:{linha_inicial}')
    for linha_acao, acao_associacao in enumerate(acoes):
        # Movendo as linhas para baixo antes de inserir os dados novos
        linha_atual = linha_inicial + (linha_acao*3)
        offset_local = linha_atual - linha_inicial
        logging.info(f'LAção:{linha_acao}, LAtual:{linha_atual}, offset:{offset_local}, Ação:{acao_associacao.acao.nome}')
        if offset_local > 0:
            insert_row(worksheet, LAST_LINE + OFFSET + offset_local, linha_atual - 1)
            insert_row(worksheet, LAST_LINE + OFFSET + offset_local, linha_atual - 1)
            insert_row(worksheet, LAST_LINE + OFFSET + offset_local, linha_atual - 1)

        sub_valores, sub_conciliacao, totalizador = sintese_receita_despesa(
                worksheet, acao_associacao, conta_associacao, periodo, fechamento_periodo, linha_atual, totalizador
            )

        total_valores += sub_valores
        total_conciliacao += sub_conciliacao

        for destinacao_idx in range(0, 2):
            if destinacao_idx == 0:
                try:
                    worksheet.unmerge_cells(f'B{linha_atual + 1}:B{linha_atual + 3}')
                except ValueError:
                    # Ignora caso o campo ja esteja desmergeado
                    pass

            row = list(worksheet.rows)[linha_atual+destinacao_idx]

            row[0].value = destinacoes[destinacao_idx]

            if destinacao_idx == 0:
                worksheet.merge_cells(f'B{linha_atual + 1}:B{linha_atual + 3}')

                thin = styles.Side(border_style="thin", color="000000")

                top_left_cell = worksheet[f'B{linha_atual + 1}']
                top_left_cell.value = acao_associacao.acao.nome
                top_left_cell.font = styles.Font(name='Arial', size=10.5, b=True, color="000000")
                top_left_cell.alignment = styles.Alignment(horizontal="left", vertical="center")

            # Colunas 15, 16 e 18 devem ter fundo cinza na linha de destinação CK (idx == 2)
            col15_cell = worksheet[f'E{linha_atual + destinacao_idx + 1}']
            col15_cell.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            col15_cell.fill = styles.PatternFill("solid", fgColor="808080" if destinacao_idx == 2 else "FFFFFF")

            col16_cell = worksheet[f'F{linha_atual + destinacao_idx + 1}']
            col16_cell.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            col16_cell.fill = styles.PatternFill("solid", fgColor="808080" if destinacao_idx == 2 else "FFFFFF")

            col18_cell = worksheet[f'I{linha_atual + destinacao_idx + 1}']
            col18_cell.border = styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            col18_cell.fill = styles.PatternFill("solid", fgColor="808080" if destinacao_idx == 2 else "FFFFFF")

    linha_atual += 3
    worksheet.merge_cells(f'B{linha_atual + 1}:B{linha_atual + 3}')

    top_left_cell = worksheet[f'B{linha_atual + 1}']
    top_left_cell.value = "TOTAL"
    top_left_cell.font = styles.Font(name='Arial', size=10.5, b=True, color="000000")
    top_left_cell.alignment = styles.Alignment(horizontal="left", vertical="center")

    for idx, destinacao in enumerate(destinacoes):
        row = list(worksheet.rows)[linha_atual + idx]
        row[CREDITO].value = formata_valor(totalizador[CREDITO][destinacao])

    OFFSET += offset_local


def bloco4_creditos_demonstrados(worksheet, receitas):
    global OFFSET

    valor_total = sum(r.valor for r in receitas)

    linha_inicial = 31 + OFFSET
    linha_atual = linha_inicial  #ACF
    offset_local = 0

    for linha, receita in enumerate(receitas):
        # Movendo as linhas para baixo antes de inserir os dados novos
        linha_atual = linha_inicial + linha
        offset_local = linha_atual - linha_inicial
        if offset_local > 0:
            insert_row(worksheet, LAST_LINE + OFFSET + offset_local, linha_atual - 1)

        row = list(worksheet.rows)[linha_atual]
        row[ITEM].value = linha + 10
        row[1].value = receita.tipo_receita.nome
        row[4].value = receita.detalhamento
        row[7].value = receita.acao_associacao.acao.nome
        row[9].value = receita.data.strftime("%d/%m/%Y")
        row[10].value = formata_valor(receita.valor)

    row = list(worksheet.rows)[linha_atual + 1]
    row[10].value = formata_valor(valor_total)

    OFFSET += offset_local


def bloco5_despesas_demonstradas(worksheet, rateios):
    pagamentos(worksheet, rateios, linha_inicial=38)


def bloco6_despesas_demonstradas(worksheet, rateios):
    pagamentos(worksheet, rateios, linha_inicial=45)


def bloco7_despesas_demonstradas(worksheet, rateios):
    pagamentos(worksheet, rateios, linha_inicial=52)


def data_geracao_documento(worksheet, usuario, previa=False):
    global OFFSET

    rows = list(worksheet.rows)
    data_geracao = date.today().strftime("%d/%m/%Y")
    tipo_texto = "parcial" if previa else "final"
    quem_gerou = "" if usuario == "" else f"pelo usuário {usuario}, "
    texto = f"Documento {tipo_texto} gerado {quem_gerou}via SIG - Escola, em: {data_geracao}"
    rows[LAST_LINE + OFFSET][0].value = texto


def sintese_receita_despesa(worksheet, acao_associacao, conta_associacao, periodo, fechamento_periodo, linha,
                            totalizador):

    logging.info(f'Resumo CUSTEIO:{acao_associacao.acao.nome} Linha:{linha}')
    row_custeio = list(worksheet.rows)[linha]
    valor_saldo_reprogramado_proximo_periodo_custeio, valor_saldo_bancario_custeio, linha, totalizador = sintese_custeio(
        row_custeio, linha, acao_associacao, conta_associacao, periodo, fechamento_periodo, totalizador
    )

    logging.info(f'Resumo CAPITAL:{acao_associacao.acao.nome} Linha:{linha}')
    row_capital = list(worksheet.rows)[linha]
    valor_saldo_reprogramado_proximo_periodo_capital, valor_saldo_bancario_capital, linha, totalizador = sintese_capital(
        row_capital, linha, acao_associacao, conta_associacao, periodo, fechamento_periodo, totalizador
    )

    logging.info(f'Resumo LIVRE:{acao_associacao.acao.nome} Linha:{linha}')
    row_livre = list(worksheet.rows)[linha]
    valor_saldo_reprogramado_proximo_periodo_livre, linha, totalizador = sintese_livre(
        row_livre, linha, valor_saldo_reprogramado_proximo_periodo_custeio,
        valor_saldo_reprogramado_proximo_periodo_capital, acao_associacao,
        conta_associacao, periodo, fechamento_periodo, totalizador
    )

    valor_total_reprogramado_proximo = valor_saldo_reprogramado_proximo_periodo_livre
    valor_total_reprogramado_proximo = (valor_total_reprogramado_proximo +
                                        valor_saldo_reprogramado_proximo_periodo_capital
                                        if valor_saldo_reprogramado_proximo_periodo_capital > 0
                                        else valor_total_reprogramado_proximo)

    valor_total_reprogramado_proximo = (valor_total_reprogramado_proximo +
                                        valor_saldo_reprogramado_proximo_periodo_custeio
                                        if valor_saldo_reprogramado_proximo_periodo_custeio > 0
                                        else valor_total_reprogramado_proximo)

    return valor_total_reprogramado_proximo, valor_saldo_bancario_capital + valor_saldo_bancario_custeio + valor_saldo_reprogramado_proximo_periodo_livre, totalizador


def sintese_custeio(row_custeio, linha, acao_associacao, conta_associacao, periodo, fechamento_periodo, totalizador):
    """
    retorna uma tupla de saldos relacionados aos custeios
    """
    saldo_reprogramado_anterior_custeio = fechamento_periodo.fechamento_anterior.saldo_reprogramado_custeio if fechamento_periodo and fechamento_periodo.fechamento_anterior else 0
    # Custeio
    receitas_demonstradas_custeio = Receita.receitas_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        categoria_receita=APLICACAO_CUSTEIO).aggregate(valor=Sum('valor'))

    rateios_demonstrados_custeio = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        aplicacao_recurso=APLICACAO_CUSTEIO).aggregate(valor=Sum('valor_rateio'))

    rateios_nao_conferidos_custeio = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False,
        aplicacao_recurso=APLICACAO_CUSTEIO).aggregate(valor=Sum('valor_rateio'))

    rateios_nao_conferidos_custeio_periodos_anteriores = RateioDespesa.rateios_da_acao_associacao_em_periodo_anteriores(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False,
        aplicacao_recurso=APLICACAO_CUSTEIO).aggregate(valor=Sum('valor_rateio'))

    valor_custeio_receitas_demonstradas = receitas_demonstradas_custeio['valor'] or 0
    valor_custeio_rateios_demonstrados = rateios_demonstrados_custeio['valor'] or 0
    valor_custeio_rateios_nao_demonstrados = rateios_nao_conferidos_custeio['valor'] or 0
    valor_custeio_rateios_nao_demonstrados_periodos_anteriores = rateios_nao_conferidos_custeio_periodos_anteriores[
                                                                     'valor'] or 0

    if saldo_reprogramado_anterior_custeio or valor_custeio_receitas_demonstradas or valor_custeio_rateios_demonstrados or valor_custeio_rateios_nao_demonstrados or valor_custeio_rateios_nao_demonstrados_periodos_anteriores:

        valor_saldo_reprogramado_proximo_periodo_custeio = (saldo_reprogramado_anterior_custeio +
                                                            valor_custeio_receitas_demonstradas -
                                                            valor_custeio_rateios_demonstrados -
                                                            valor_custeio_rateios_nao_demonstrados)

        valor_saldo_bancario_custeio = (valor_saldo_reprogramado_proximo_periodo_custeio +
                                        valor_custeio_rateios_nao_demonstrados +
                                        valor_custeio_rateios_nao_demonstrados_periodos_anteriores)

        valor_saldo_bancario_custeio = valor_saldo_bancario_custeio if valor_saldo_bancario_custeio > 0 else 0

        logging.info(f'CRÉDITO CUSTEIO:{acao_associacao.acao.nome} > {valor_custeio_receitas_demonstradas}')

        row_custeio[SALDO_ANTERIOR].value = formata_valor(saldo_reprogramado_anterior_custeio)
        row_custeio[CREDITO].value = formata_valor(valor_custeio_receitas_demonstradas)
        row_custeio[DESPESA_REALIZADA].value = formata_valor(valor_custeio_rateios_demonstrados)
        row_custeio[DESPESA_NAO_REALIZADA].value = formata_valor(valor_custeio_rateios_nao_demonstrados)

        row_custeio[SALDO_REPROGRAMADO_PROXIMO].value = formata_valor(
            valor_saldo_reprogramado_proximo_periodo_custeio if valor_saldo_reprogramado_proximo_periodo_custeio > 0
            else 0)

        row_custeio[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].value = formata_valor(
            valor_custeio_rateios_nao_demonstrados_periodos_anteriores)

        row_custeio[SALDO_BANCARIO].value = formata_valor(valor_saldo_bancario_custeio)

        linha += 1

        totalizador[CREDITO]['C'] += valor_custeio_receitas_demonstradas

        return valor_saldo_reprogramado_proximo_periodo_custeio, valor_saldo_bancario_custeio, linha, totalizador

    else:
        logging.info(f'CRÉDITO CUSTEIO:{acao_associacao.acao.nome} > ZERO')
        row_custeio[SALDO_ANTERIOR].value = formata_valor(0)
        row_custeio[CREDITO].value = formata_valor(0)
        row_custeio[DESPESA_REALIZADA].value = formata_valor(0)
        row_custeio[DESPESA_NAO_REALIZADA].value = formata_valor(0)

        row_custeio[SALDO_REPROGRAMADO_PROXIMO].value = formata_valor(0)

        row_custeio[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].value = formata_valor(0)

        row_custeio[SALDO_BANCARIO].value = formata_valor(0)

    linha += 1

    return 0, 0, linha, totalizador


def sintese_capital(row_capital, linha, acao_associacao, conta_associacao, periodo, fechamento_periodo, totalizador):
    """
    retorna uma tupla de saldos relacionados aos capitais
    """
    saldo_reprogramado_anterior_capital = (fechamento_periodo.fechamento_anterior.saldo_reprogramado_capital
                                           if fechamento_periodo and fechamento_periodo.fechamento_anterior else 0)

    receitas_demonstradas_capital = Receita.receitas_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        categoria_receita=APLICACAO_CAPITAL).aggregate(valor=Sum('valor'))

    rateios_demonstrados_capital = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        aplicacao_recurso=APLICACAO_CAPITAL).aggregate(valor=Sum('valor_rateio'))

    rateios_nao_conferidos_capital = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False,
        aplicacao_recurso=APLICACAO_CAPITAL).aggregate(valor=Sum('valor_rateio'))

    rateios_nao_conferidos_outros_periodos = RateioDespesa.rateios_da_acao_associacao_em_periodo_anteriores(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False,
        aplicacao_recurso=APLICACAO_CAPITAL).aggregate(valor=Sum('valor_rateio'))

    valor_capital_receitas_demonstradas = receitas_demonstradas_capital['valor'] or 0
    valor_capital_rateios_demonstrados = rateios_demonstrados_capital['valor'] or 0
    valor_capital_rateios_nao_demonstrados = rateios_nao_conferidos_capital['valor'] or 0
    valor_capital_rateios_nao_demonstrados_outros_periodos = rateios_nao_conferidos_outros_periodos['valor'] or 0

    if (
        saldo_reprogramado_anterior_capital or
        valor_capital_receitas_demonstradas or
        valor_capital_rateios_demonstrados or
        valor_capital_rateios_nao_demonstrados or
        valor_capital_rateios_nao_demonstrados_outros_periodos
    ):

        valor_saldo_reprogramado_proximo_periodo_capital = saldo_reprogramado_anterior_capital + \
                                                           valor_capital_receitas_demonstradas - \
                                                           valor_capital_rateios_demonstrados - \
                                                           valor_capital_rateios_nao_demonstrados

        valor_saldo_bancario_capital = valor_saldo_reprogramado_proximo_periodo_capital + valor_capital_rateios_nao_demonstrados + valor_capital_rateios_nao_demonstrados_outros_periodos
        valor_saldo_bancario_capital = valor_saldo_bancario_capital if valor_saldo_bancario_capital > 0 else 0

        logging.info(f'CRÉDITO CAPITAL:{acao_associacao.acao.nome} > {valor_capital_receitas_demonstradas}')

        row_capital[SALDO_ANTERIOR].value = formata_valor(saldo_reprogramado_anterior_capital)
        row_capital[CREDITO].value = formata_valor(valor_capital_receitas_demonstradas)
        row_capital[DESPESA_REALIZADA].value = formata_valor(valor_capital_rateios_demonstrados)
        row_capital[DESPESA_NAO_REALIZADA].value = formata_valor(valor_capital_rateios_nao_demonstrados)

        row_capital[SALDO_REPROGRAMADO_PROXIMO].value = formata_valor(
            valor_saldo_reprogramado_proximo_periodo_capital
            if valor_saldo_reprogramado_proximo_periodo_capital > 0 else 0)
        row_capital[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].value = formata_valor(valor_capital_rateios_nao_demonstrados_outros_periodos)
        row_capital[SALDO_BANCARIO].value = formata_valor(valor_saldo_bancario_capital)

        linha += 1

        totalizador[CREDITO]['K'] += valor_capital_receitas_demonstradas

        return valor_saldo_reprogramado_proximo_periodo_capital, valor_saldo_bancario_capital, linha, totalizador

    else:
        logging.info(f'CRÉDITO CAPITAL:{acao_associacao.acao.nome} > ZERO')

        row_capital[SALDO_ANTERIOR].value = formata_valor(0)
        row_capital[CREDITO].value = formata_valor(0)
        row_capital[DESPESA_REALIZADA].value = formata_valor(0)
        row_capital[DESPESA_NAO_REALIZADA].value = formata_valor(0)

        row_capital[SALDO_REPROGRAMADO_PROXIMO].value = formata_valor(0)
        row_capital[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].value = formata_valor(0)
        row_capital[SALDO_BANCARIO].value = formata_valor(0)

    linha += 1

    return 0, 0, linha, totalizador


def sintese_livre(row_livre, linha, valor_saldo_reprogramado_proximo_periodo_custeio,
                  valor_saldo_reprogramado_proximo_periodo_capital, acao_associacao,
                  conta_associacao, periodo, fechamento_periodo, totalizador):

    saldo_reprogramado_anterior_livre = (fechamento_periodo.fechamento_anterior.saldo_reprogramado_livre
                                         if fechamento_periodo and fechamento_periodo.fechamento_anterior else 0)

    receitas_demonstradas_livre = Receita.receitas_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        categoria_receita=APLICACAO_LIVRE).aggregate(valor=Sum('valor'))

    valor_livre_receitas_demonstradas = receitas_demonstradas_livre['valor'] or 0

    if (
        saldo_reprogramado_anterior_livre or
        valor_livre_receitas_demonstradas or
        valor_saldo_reprogramado_proximo_periodo_custeio < 0 or
        valor_saldo_reprogramado_proximo_periodo_capital < 0
    ):

        valor_saldo_reprogramado_proximo_periodo_livre = (saldo_reprogramado_anterior_livre +
                                                          valor_livre_receitas_demonstradas)

        valor_saldo_reprogramado_proximo_periodo_livre = (valor_saldo_reprogramado_proximo_periodo_livre +
                                                          valor_saldo_reprogramado_proximo_periodo_capital
                                                          if valor_saldo_reprogramado_proximo_periodo_capital < 0
                                                          else valor_saldo_reprogramado_proximo_periodo_livre)

        valor_saldo_reprogramado_proximo_periodo_livre = (valor_saldo_reprogramado_proximo_periodo_livre +
                                                          valor_saldo_reprogramado_proximo_periodo_custeio
                                                          if valor_saldo_reprogramado_proximo_periodo_custeio < 0
                                                          else valor_saldo_reprogramado_proximo_periodo_livre)

        logging.info(f'CRÉDITO LIVRE:{acao_associacao.acao.nome} > {valor_livre_receitas_demonstradas}')

        cor_cinza = styles.colors.Color(rgb='808080')
        fill = styles.fills.PatternFill(patternType='solid', fgColor=cor_cinza)
        row_livre[DESPESA_REALIZADA].fill = fill
        row_livre[DESPESA_NAO_REALIZADA].fill = fill
        row_livre[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].fill = fill

        row_livre[SALDO_ANTERIOR].value = formata_valor(saldo_reprogramado_anterior_livre)
        row_livre[CREDITO].value = formata_valor(valor_livre_receitas_demonstradas)
        row_livre[
            SALDO_REPROGRAMADO_PROXIMO].value = formata_valor(valor_saldo_reprogramado_proximo_periodo_livre)

        linha += 1

        totalizador[CREDITO]['CK'] += valor_livre_receitas_demonstradas

        return valor_saldo_reprogramado_proximo_periodo_livre, linha, totalizador

    else:
        logging.info(f'CRÉDITO LIVRE:{acao_associacao.acao.nome} > ZERO')

        cor_cinza = styles.colors.Color(rgb='808080')
        fill = styles.fills.PatternFill(patternType='solid', fgColor=cor_cinza)
        row_livre[DESPESA_REALIZADA].fill = fill
        row_livre[DESPESA_NAO_REALIZADA].fill = fill
        row_livre[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].fill = fill

        row_livre[SALDO_ANTERIOR].value = formata_valor(0)
        row_livre[CREDITO].value = formata_valor(0)
        row_livre[SALDO_REPROGRAMADO_PROXIMO].value = formata_valor(0)

    linha += 1

    return 0, linha, totalizador


def pagamentos(worksheet, rateios, linha_inicial=0):
    global OFFSET

    valor_total = sum(r.valor_rateio for r in rateios)

    linha_inicial += OFFSET
    linha_atual = linha_inicial
    offsetlocal = 0

    for linha, rateio in enumerate(rateios):
        linha_atual = linha_inicial + linha
        offsetlocal = linha_atual - linha_inicial
        if offsetlocal > 0:
            insert_row(worksheet, LAST_LINE + OFFSET + offsetlocal, linha_atual - 1)

        row = list(worksheet.rows)[linha_atual]
        row[ITEM].value = linha + 1
        row[RAZAO_SOCIAL].value = rateio.despesa.nome_fornecedor
        row[CNPJ_CPF].value = rateio.despesa.cpf_cnpj_fornecedor
        row[TIPO_DOCUMENTO].value = \
            rateio.despesa.tipo_documento.nome if rateio.despesa.tipo_documento else ''
        row[NUMERO_DOCUMENTO].value = rateio.despesa.numero_documento
        row[ACAO_DOCUMENTO].value = rateio.acao_associacao.acao.nome
        row[ESPECIFICACAO_MATERIAL].value = \
            rateio.especificacao_material_servico.descricao if rateio.especificacao_material_servico else ''
        row[TIPO_DESPESA].value = rateio.aplicacao_recurso

        tipo_transacao = ''
        if rateio.conta_associacao.tipo_conta:
            if rateio.conta_associacao.tipo_conta.nome == 'Cheque':
                tipo_transacao = rateio.despesa.documento_transacao
            else:
                tipo_transacao = rateio.despesa.tipo_transacao.nome

        row[TIPO_TRANSACAO].value = tipo_transacao
        row[DATA].value = rateio.despesa.data_documento.strftime("%d/%m/%Y") if rateio.despesa.data_documento else ''
        row[DATA_2].value = rateio.despesa.data_documento.strftime("%d/%m/%Y") if rateio.despesa.data_documento else ''
        row[VALOR].value = formata_valor(rateio.valor_rateio)

    row = list(worksheet.rows)[linha_atual + 1]
    row[11].value = formata_valor(valor_total)

    OFFSET += offsetlocal


def observacoes(worksheet, acao_associacao, periodo, conta_associacao):
    #precisa mudar ainda. pois aqui é por acao...

    start_line = 44
    row = list(worksheet.rows)[start_line]
    row[ITEM].value = ObservacaoConciliacao.objects.filter(
        acao_associacao=acao_associacao,
        conta_associacao=conta_associacao,
        periodo__uuid=periodo.uuid
    ).first().texto if ObservacaoConciliacao.objects.filter(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo__uuid=periodo.uuid
    ).exists() else ''


def formata_valor(valor):
    from babel.numbers import format_currency
    sinal, valor_formatado = format_currency(valor, 'BRL', locale='pt_BR').split('\xa0')
    sinal = '-' if '-' in sinal else ''
    return f'{sinal}{valor_formatado}'
