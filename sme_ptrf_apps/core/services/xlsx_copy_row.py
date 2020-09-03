import re
from copy import copy

from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries


def copy_row(ws, source_row, dest_row, copy_data=False, copy_style=True, copy_merged_columns=True):
    """Copia uma linha da planilha para a linha imediatamento abaixo mantendo todos os atributos."""
    CELL_RE = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    def replace(m):
        row = m.group('row')
        prefix = "$" if row.find("$") != -1 else ""
        row = int(row.replace("$", ""))
        row += dest_row if row > source_row else 0
        return m.group('col') + prefix + str(row)

    # Fazendo unmerge das celulas de destino.
    mergedcells = []
    for group in ws.merged_cells.ranges:
        mergedcells.append(group)

    for cr in mergedcells:
        min_col, min_row, max_col, max_row = range_boundaries(str(cr))
        if max_row == min_row == source_row + 1:
            ws.unmerge_cells(str(cr))

    new_row_idx = source_row + 1
    for row in range(new_row_idx, new_row_idx + dest_row):
        new_rd = copy(ws.row_dimensions[row - 1])
        new_rd.index = row
        ws.row_dimensions[row] = new_rd

        for col in range(ws.max_column):
            col = get_column_letter(col + 1)
            cell = ws['%s%d' % (col, row)]
            source = ws['%s%d' % (col, row - 1)]
            if copy_style:
                cell._style = copy(source._style)
            if copy_data and not isinstance(cell, MergedCell):
                cell.data_type = source.data_type
                cell.value = source.value

    for cr_idx, cr in enumerate(ws.merged_cell_ranges):
        ws.merged_cell_ranges[cr_idx] = CELL_RE.sub(
            replace,
            str(cr)
        )

    if copy_merged_columns:
        for cr in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cr))
            if max_row == min_row == source_row:
                for row in range(new_row_idx, new_row_idx + dest_row):
                    newCellRange = get_column_letter(min_col) + str(row) + ":" + get_column_letter(max_col) + str(row)
                    ws.merge_cells(newCellRange)
