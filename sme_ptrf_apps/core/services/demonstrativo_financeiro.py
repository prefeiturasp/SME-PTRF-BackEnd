import logging
import os

import re
from copy import copy
from tempfile import NamedTemporaryFile

from datetime import date


from django.db.models import Sum
from django.contrib.staticfiles.storage import staticfiles_storage
from django.core.files import File
from openpyxl import load_workbook, styles
from sme_ptrf_apps.core.services.xlsx_copy_row import copy_row

from sme_ptrf_apps.core.choices import MembroEnum

from sme_ptrf_apps.core.models import (FechamentoPeriodo, MembroAssociacao, ObservacaoConciliacao,
                                       DemonstrativoFinanceiro)
from sme_ptrf_apps.despesas.models import RateioDespesa
from sme_ptrf_apps.receitas.models import Receita
from sme_ptrf_apps.receitas.tipos_aplicacao_recurso_receitas import (APLICACAO_CAPITAL, APLICACAO_CUSTEIO,
                                                                     APLICACAO_LIVRE)


LOGGER = logging.getLogger(__name__)

COL_CABECALHO = 9
LINHA_PERIODO_CABECALHO = 4
LINHA_ACAO_CABECALHO = 5
LINHA_CONTA_CABECALHO = 6
LAST_LINE = 56

# Coluna 2 da planilha
SALDO_ANTERIOR = 0
CREDITO = 2
DESPESA_REALIZADA = 3
DESPESA_NAO_REALIZADA = 4
SALDO_REPROGRAMADO_PROXIMO = 5
TOTAL_REPROGRAMADO_PROXIMO = 6
DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS = 7
SALDO_BANCARIO = 8
TOTAL_SALDO_BANCARIO = 10

# Colunas dos blocos 4 e 5 da planilha
ITEM = 0
RAZAO_SOCIAL = 1
CNPJ_CPF = 2
TIPO_DOCUMENTO = 3
NUMERO_DOCUMENTO = 4
DATA = 5
ESPECIFICACAO_MATERIAL = 6
TIPO_DESPESA = 7
TIPO_TRANSACAO = 8
DATA_2 = 9
VALOR = 10


def gerar_arquivo_demonstrativo_financeiro(periodo, acao_associacao, conta_associacao, prestacao):
    filename = 'demonstrativo_financeiro.xlsx'

    xlsx = gerar(periodo, acao_associacao, conta_associacao)

    with NamedTemporaryFile() as tmp:
        xlsx.save(tmp.name)

        demonstrativo_financeiro, _ = DemonstrativoFinanceiro.objects.update_or_create(
            acao_associacao=acao_associacao,
            conta_associacao=conta_associacao,
            prestacao_conta=prestacao
        )
        demonstrativo_financeiro.arquivo.save(name=filename, content=File(tmp))


def gerar(periodo, acao_associacao, conta_associacao, previa=False):

    LOGGER.info("GERANDO DEMONSTRATIVO...")
    rateios_conferidos = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True)

    rateios_nao_conferidos = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False)

    rateios_nao_conferidos_em_periodos_anteriores = RateioDespesa.rateios_da_acao_associacao_em_periodo_anteriores(
        acao_associacao=acao_associacao, periodo=periodo, conta_associacao=conta_associacao, conferido=False)

    receitas_demonstradas = Receita.receitas_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True)

    fechamento_periodo = FechamentoPeriodo.objects.filter(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo__uuid=periodo.uuid).first()

    path = os.path.join(os.path.basename(staticfiles_storage.location), 'cargas')
    nome_arquivo = os.path.join(path, 'modelo_demonstrativo_financeiro.xlsx')
    workbook = load_workbook(nome_arquivo)
    worksheet = workbook.active
    try:
        cabecalho(worksheet, periodo, acao_associacao, conta_associacao, previa)
        identificacao_apm(worksheet, acao_associacao)
        observacoes(worksheet, acao_associacao, periodo, conta_associacao)
        data_geracao_documento(worksheet, previa)
        sintese_receita_despesa(worksheet, acao_associacao, conta_associacao, periodo, fechamento_periodo)
        creditos_demonstrados(worksheet, receitas_demonstradas)
        acc = len(receitas_demonstradas) - 1 if len(receitas_demonstradas) > 1 else 0
        pagamentos(worksheet, rateios_conferidos, acc=acc, start_line=28)
        acc += len(rateios_conferidos) - 1 if len(rateios_conferidos) > 1 else 0
        pagamentos(worksheet, rateios_nao_conferidos, acc=acc, start_line=34)
        acc += len(rateios_nao_conferidos) - 1 if len(rateios_nao_conferidos) > 1 else 0
        pagamentos(worksheet, rateios_nao_conferidos_em_periodos_anteriores, acc=acc, start_line=40)
    except Exception as e:
        LOGGER.info("ERRO no Demonstrativo: %s", str(e))

    LOGGER.info("DEMONSTRATIVO GERADO")
    return workbook


def cabecalho(worksheet, periodo, acao_associacao, conta_associacao, previa):
    rows = list(worksheet.rows)
    texto = "Demonstrativo Financeiro - PRÉVIA" if previa else "Demonstrativo Financeiro - FINAL"
    rows[2][6].value = texto
    rows[LINHA_PERIODO_CABECALHO][COL_CABECALHO].value = str(periodo)
    rows[LINHA_ACAO_CABECALHO][COL_CABECALHO].value = acao_associacao.acao.nome
    rows[LINHA_CONTA_CABECALHO][COL_CABECALHO].value = conta_associacao.tipo_conta.nome


def identificacao_apm(worksheet, acao_associacao):
    """BLOCO 1 - IDENTIFICAÇÃO DA APM/APMSUAC DA UNIDADE EDUCACIONAL"""
    associacao = acao_associacao.associacao
    rows = list(worksheet.rows)
    rows[10][0].value = associacao.nome
    rows[10][6].value = associacao.cnpj
    rows[10][7].value = associacao.unidade.codigo_eol or ""
    rows[10][8].value = associacao.unidade.dre.nome if associacao.unidade.dre else ""

    presidente_diretoria_executiva = MembroAssociacao.objects.filter(associacao=associacao,
                                                                     cargo_associacao=MembroEnum.PRESIDENTE_DIRETORIA_EXECUTIVA.name).first()

    presidente_conselho_fiscal = MembroAssociacao.objects.filter(associacao=associacao,
                                                                 cargo_associacao=MembroEnum.PRESIDENTE_CONSELHO_FISCAL.name).first()

    rows[LAST_LINE - 4][0].value = presidente_diretoria_executiva.nome if presidente_diretoria_executiva else ''
    rows[LAST_LINE - 4][6].value = presidente_conselho_fiscal.nome if presidente_conselho_fiscal else ''


def data_geracao_documento(worksheet, previa=False):
    rows = list(worksheet.rows)
    data_geracao = date.today().strftime("%d/%m/%Y")
    texto = f"Documento parcial gerado em: {data_geracao}" if previa else f"Documento final gerado em: {data_geracao}"
    rows[LAST_LINE - 1][0].value = texto


def sintese_receita_despesa(worksheet, acao_associacao, conta_associacao, periodo, fechamento_periodo):
    linha = 15
    valor_saldo_reprogramado_proximo_periodo_custeio = 0
    valor_saldo_bancario_custeio = 0
    valor_saldo_reprogramado_proximo_periodo_capital = 0
    valor_saldo_bancario_capital = 0
    valor_saldo_reprogramado_proximo_periodo_livre = 0

    row_custeio = list(worksheet.rows)[linha]
    valor_saldo_reprogramado_proximo_periodo_custeio, valor_saldo_bancario_custeio, linha = sintese_custeio(
        row_custeio, linha, acao_associacao, conta_associacao, periodo, fechamento_periodo
    )

    row_capital = list(worksheet.rows)[linha]
    valor_saldo_reprogramado_proximo_periodo_capital, valor_saldo_bancario_capital, linha = sintese_capital(
        row_capital, linha, acao_associacao, conta_associacao, periodo, fechamento_periodo
    )

    row_livre = list(worksheet.rows)[linha]
    valor_saldo_reprogramado_proximo_periodo_livre, linha = sintese_livre(
        row_livre, linha, valor_saldo_reprogramado_proximo_periodo_custeio,
        valor_saldo_reprogramado_proximo_periodo_capital, acao_associacao,
        conta_associacao, periodo, fechamento_periodo
    )

    valor_total_reprogramado_proximo = valor_saldo_reprogramado_proximo_periodo_livre
    valor_total_reprogramado_proximo = valor_total_reprogramado_proximo + \
        valor_saldo_reprogramado_proximo_periodo_capital if valor_saldo_reprogramado_proximo_periodo_capital > 0 else valor_total_reprogramado_proximo
    valor_total_reprogramado_proximo = valor_total_reprogramado_proximo + \
        valor_saldo_reprogramado_proximo_periodo_custeio if valor_saldo_reprogramado_proximo_periodo_custeio > 0 else valor_total_reprogramado_proximo
    row_livre[
        SALDO_BANCARIO].value = f'L {formata_valor(valor_saldo_reprogramado_proximo_periodo_livre)}' if valor_saldo_reprogramado_proximo_periodo_livre != 0 else ''
    row_custeio[TOTAL_REPROGRAMADO_PROXIMO].value = formata_valor(valor_total_reprogramado_proximo)
    row_custeio[TOTAL_SALDO_BANCARIO].value = formata_valor(
        valor_saldo_bancario_capital + valor_saldo_bancario_custeio + valor_saldo_reprogramado_proximo_periodo_livre)


def sintese_custeio(row_custeio, linha, acao_associacao, conta_associacao, periodo, fechamento_periodo):
    """
    retorna uma tupla de saldos relacionados aos custeios
    """
    saldo_reprogramado_anterior_custeio = fechamento_periodo.fechamento_anterior.saldo_reprogramado_custeio if fechamento_periodo and fechamento_periodo.fechamento_anterior else 0
    # Custeio
    receitas_demonstradas_custeio = Receita.receitas_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        categoria_receita=APLICACAO_CUSTEIO).aggregate(valor=Sum('valor'))

    rateios_demonstrados_custeio = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        aplicacao_recurso=APLICACAO_CUSTEIO).aggregate(valor=Sum('valor_rateio'))

    rateios_nao_conferidos_custeio = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False,
        aplicacao_recurso=APLICACAO_CUSTEIO).aggregate(valor=Sum('valor_rateio'))

    rateios_nao_conferidos_custeio_periodos_anteriores = RateioDespesa.rateios_da_acao_associacao_em_periodo_anteriores(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False,
        aplicacao_recurso=APLICACAO_CUSTEIO).aggregate(valor=Sum('valor_rateio'))

    valor_custeio_receitas_demonstradas = receitas_demonstradas_custeio['valor'] or 0
    valor_custeio_rateios_demonstrados = rateios_demonstrados_custeio['valor'] or 0
    valor_custeio_rateios_nao_demonstrados = rateios_nao_conferidos_custeio['valor'] or 0
    valor_custeio_rateios_nao_demonstrados_periodos_anteriores = rateios_nao_conferidos_custeio_periodos_anteriores[
        'valor'] or 0

    if saldo_reprogramado_anterior_custeio or valor_custeio_receitas_demonstradas or valor_custeio_rateios_demonstrados or valor_custeio_rateios_nao_demonstrados or valor_custeio_rateios_nao_demonstrados_periodos_anteriores:
        row_custeio[SALDO_ANTERIOR].value = f'C {formata_valor(saldo_reprogramado_anterior_custeio)}'
        row_custeio[CREDITO].value = f'C {formata_valor(valor_custeio_receitas_demonstradas)}'
        row_custeio[DESPESA_REALIZADA].value = f'C {formata_valor(valor_custeio_rateios_demonstrados)}'
        row_custeio[DESPESA_NAO_REALIZADA].value = f'C {formata_valor(valor_custeio_rateios_nao_demonstrados)}'
        valor_saldo_reprogramado_proximo_periodo_custeio = saldo_reprogramado_anterior_custeio + \
            valor_custeio_receitas_demonstradas - \
            valor_custeio_rateios_demonstrados - \
            valor_custeio_rateios_nao_demonstrados
        row_custeio[
            SALDO_REPROGRAMADO_PROXIMO].value = f'C {formata_valor(valor_saldo_reprogramado_proximo_periodo_custeio if valor_saldo_reprogramado_proximo_periodo_custeio > 0 else 0)}'

        row_custeio[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].value = f'C {formata_valor(valor_custeio_rateios_nao_demonstrados_periodos_anteriores)}'
        valor_saldo_bancario_custeio = valor_saldo_reprogramado_proximo_periodo_custeio + valor_custeio_rateios_nao_demonstrados
        valor_saldo_bancario_custeio = valor_saldo_bancario_custeio if valor_saldo_bancario_custeio > 0 else 0
        row_custeio[SALDO_BANCARIO].value = f'C {formata_valor(valor_saldo_bancario_custeio)}'
        linha += 1

        return valor_saldo_reprogramado_proximo_periodo_custeio, valor_saldo_bancario_custeio, linha

    return (0, 0, linha)


def sintese_capital(row_capital, linha, acao_associacao, conta_associacao, periodo, fechamento_periodo):
    """
    retorna uma tupla de saldos relacionados aos capitais
    """
    saldo_reprogramado_anterior_capital = fechamento_periodo.fechamento_anterior.saldo_reprogramado_capital if fechamento_periodo and fechamento_periodo.fechamento_anterior else 0
    receitas_demonstradas_capital = Receita.receitas_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        categoria_receita=APLICACAO_CAPITAL).aggregate(valor=Sum('valor'))

    rateios_demonstrados_capital = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        aplicacao_recurso=APLICACAO_CAPITAL).aggregate(valor=Sum('valor_rateio'))

    rateios_nao_conferidos_capital = RateioDespesa.rateios_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False,
        aplicacao_recurso=APLICACAO_CAPITAL).aggregate(valor=Sum('valor_rateio'))

    rateios_nao_conferidos_outros_periodos = RateioDespesa.rateios_da_acao_associacao_em_periodo_anteriores(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=False,
        aplicacao_recurso=APLICACAO_CAPITAL).aggregate(valor=Sum('valor_rateio'))

    valor_capital_receitas_demonstradas = receitas_demonstradas_capital['valor'] or 0
    valor_capital_rateios_demonstrados = rateios_demonstrados_capital['valor'] or 0
    valor_capital_rateios_nao_demonstrados = rateios_nao_conferidos_capital['valor'] or 0
    valor_capital_rateios_nao_demonstrados_outros_periodos = rateios_nao_conferidos_outros_periodos['valor'] or 0

    if saldo_reprogramado_anterior_capital or valor_capital_receitas_demonstradas or valor_capital_rateios_demonstrados or valor_capital_rateios_nao_demonstrados or valor_capital_rateios_nao_demonstrados_outros_periodos:
        row_capital[SALDO_ANTERIOR].value = f'K {formata_valor(saldo_reprogramado_anterior_capital)}'
        row_capital[CREDITO].value = f'K {formata_valor(valor_capital_receitas_demonstradas)}'
        row_capital[DESPESA_REALIZADA].value = f'K {formata_valor(valor_capital_rateios_demonstrados)}'
        row_capital[DESPESA_NAO_REALIZADA].value = f'K {formata_valor(valor_capital_rateios_nao_demonstrados)}'
        valor_saldo_reprogramado_proximo_periodo_capital = saldo_reprogramado_anterior_capital + \
            valor_capital_receitas_demonstradas - \
            valor_capital_rateios_demonstrados - \
            valor_capital_rateios_nao_demonstrados
        row_capital[SALDO_REPROGRAMADO_PROXIMO].value = f'K {formata_valor(valor_saldo_reprogramado_proximo_periodo_capital if valor_saldo_reprogramado_proximo_periodo_capital > 0 else 0)}'
        row_capital[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].value = f'K {formata_valor(valor_capital_rateios_nao_demonstrados_outros_periodos)}'
        valor_saldo_bancario_capital = valor_saldo_reprogramado_proximo_periodo_capital + valor_capital_rateios_nao_demonstrados
        valor_saldo_bancario_capital = valor_saldo_bancario_capital if valor_saldo_bancario_capital > 0 else 0
        row_capital[SALDO_BANCARIO].value = f'K {formata_valor(valor_saldo_bancario_capital)}'
        linha += 1

        return valor_saldo_reprogramado_proximo_periodo_capital, valor_saldo_bancario_capital, linha

    return (0, 0, linha)


def sintese_livre(row_livre, linha, valor_saldo_reprogramado_proximo_periodo_custeio,
                  valor_saldo_reprogramado_proximo_periodo_capital, acao_associacao,
                  conta_associacao, periodo, fechamento_periodo):
    saldo_reprogramado_anterior_livre = fechamento_periodo.fechamento_anterior.saldo_reprogramado_livre if fechamento_periodo and fechamento_periodo.fechamento_anterior else 0
    receitas_demonstradas_livre = Receita.receitas_da_acao_associacao_no_periodo(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo=periodo, conferido=True,
        categoria_receita=APLICACAO_LIVRE).aggregate(valor=Sum('valor'))

    valor_livre_receitas_demonstradas = receitas_demonstradas_livre['valor'] or 0

    if saldo_reprogramado_anterior_livre or valor_livre_receitas_demonstradas or valor_saldo_reprogramado_proximo_periodo_custeio < 0 or valor_saldo_reprogramado_proximo_periodo_capital < 0:
        row_livre[SALDO_ANTERIOR].value = f'L {formata_valor(saldo_reprogramado_anterior_livre)}'
        row_livre[CREDITO].value = f'L {formata_valor(valor_livre_receitas_demonstradas)}'
        cor_cinza = styles.colors.Color(rgb='808080')
        fill = styles.fills.PatternFill(patternType='solid', fgColor=cor_cinza)
        row_livre[DESPESA_REALIZADA].fill = fill
        row_livre[DESPESA_NAO_REALIZADA].fill = fill
        row_livre[DESPESA_NAO_DEMONSTRADA_OUTROS_PERIODOS].fill = fill
        valor_saldo_reprogramado_proximo_periodo_livre = saldo_reprogramado_anterior_livre + valor_livre_receitas_demonstradas
        valor_saldo_reprogramado_proximo_periodo_livre = valor_saldo_reprogramado_proximo_periodo_livre + \
            valor_saldo_reprogramado_proximo_periodo_capital if valor_saldo_reprogramado_proximo_periodo_capital < 0 else valor_saldo_reprogramado_proximo_periodo_livre
        valor_saldo_reprogramado_proximo_periodo_livre = valor_saldo_reprogramado_proximo_periodo_livre + \
            valor_saldo_reprogramado_proximo_periodo_custeio if valor_saldo_reprogramado_proximo_periodo_custeio < 0 else valor_saldo_reprogramado_proximo_periodo_livre
        row_livre[
            SALDO_REPROGRAMADO_PROXIMO].value = f'L {formata_valor(valor_saldo_reprogramado_proximo_periodo_livre)}'

        return valor_saldo_reprogramado_proximo_periodo_livre, linha

    return (0, linha)


def creditos_demonstrados(worksheet, receitas, acc=0, start_line=22):
    quantidade = acc
    last_line = LAST_LINE + quantidade
    valor_total = sum(r.valor for r in receitas)
    ind = start_line

    for linha, receita in enumerate(receitas):
        # Movendo as linhas para baixo antes de inserir os dados novos
        ind = start_line + quantidade + linha
        if linha > 0:
            for row_idx in range(last_line + linha, ind - 2, -1):
                copy_row(worksheet, row_idx, 1, copy_data=True)

        row = list(worksheet.rows)[ind - 1]
        row[ITEM].value = linha + 1
        row[1].value = receita.tipo_receita.nome
        row[4].value = receita.detalhamento
        row[7].value = receita.data.strftime("%d/%m/%Y")
        row[9].value = formata_valor(receita.valor)

    row = list(worksheet.rows)[(ind - 1) + 1]
    row[9].value = formata_valor(valor_total)


def pagamentos(worksheet, rateios, acc=0, start_line=26):
    """
    BLOCO 4 - DESPESAS EFETUADAS NO PERÍODO
    BLOCO 5 - DESPESAS EFETUADAS E NÃO DEMONSTRADAS NO EXTRATO DO PERÍODO (ATUAL E ANTERIORES)
    """

    quantidade = acc if acc else 0
    last_line = LAST_LINE + quantidade
    valor_total = sum(r.valor_rateio for r in rateios)
    ind = start_line + quantidade
    for linha, rateio in enumerate(rateios):
        # Movendo as linhas para baixo antes de inserir os dados novos
        ind = start_line + quantidade + linha
        if linha > 0:
            for row_idx in range(last_line + linha, ind - 2, -1):
                copy_row(worksheet, row_idx, 1, copy_data=True)

        row = list(worksheet.rows)[ind - 1]
        row[ITEM].value = linha + 1
        row[RAZAO_SOCIAL].value = rateio.despesa.nome_fornecedor
        row[CNPJ_CPF].value = rateio.despesa.cpf_cnpj_fornecedor
        row[TIPO_DOCUMENTO].value = rateio.despesa.tipo_documento.nome if rateio.despesa.tipo_documento else ''
        row[NUMERO_DOCUMENTO].value = rateio.despesa.numero_documento
        row[
            ESPECIFICACAO_MATERIAL].value = rateio.especificacao_material_servico.descricao if rateio.especificacao_material_servico else ''
        row[TIPO_DESPESA].value = rateio.aplicacao_recurso

        tipo_transacao = ''
        if rateio.conta_associacao.tipo_conta:
            if rateio.conta_associacao.tipo_conta.nome == 'Cheque':
                tipo_transacao = rateio.despesa.documento_transacao
            else:
                tipo_transacao = rateio.despesa.tipo_transacao.nome

        row[TIPO_TRANSACAO].value = tipo_transacao
        row[DATA].value = rateio.despesa.data_documento.strftime("%d/%m/%Y") if rateio.despesa.data_documento else ''
        row[DATA_2].value = rateio.despesa.data_documento.strftime("%d/%m/%Y") if rateio.despesa.data_documento else ''
        row[VALOR].value = formata_valor(rateio.valor_rateio)

    row = list(worksheet.rows)[(ind - 1) + 1]
    row[9].value = formata_valor(valor_total)


def observacoes(worksheet, acao_associacao, periodo, conta_associacao):
    """BLOCO 6 - OBSERVAÇÃO"""

    start_line = 44
    row = list(worksheet.rows)[start_line]
    row[ITEM].value = ObservacaoConciliacao.objects.filter(
        acao_associacao=acao_associacao,
        conta_associacao=conta_associacao,
        periodo__uuid=periodo.uuid
    ).first().texto if ObservacaoConciliacao.objects.filter(
        acao_associacao=acao_associacao, conta_associacao=conta_associacao, periodo__uuid=periodo.uuid
    ).exists() else ''


def formata_valor(valor):
    from babel.numbers import format_currency
    sinal, valor_formatado = format_currency(valor, 'BRL', locale='pt_BR').split('\xa0')
    sinal = '-' if '-' in sinal else ''
    return f'{sinal}{valor_formatado}'
