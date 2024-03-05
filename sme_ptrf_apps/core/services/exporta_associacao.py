import logging
import os

from django.contrib.staticfiles.storage import staticfiles_storage
from openpyxl import load_workbook

from ..choices.membro_associacao import RepresentacaoCargo
from waffle import get_waffle_flag_model

LOGGER = logging.getLogger(__name__)

# Worksheets
ASSOCIACAO = 0
MEMBROS = 1
CONTAS = 2

# Linhas Dados Básicos
NOME_ASSOCIACAO = 0
EOL = 1
DRE = 2
CNPJ = 3
CCM = 4
EMAIL_ASSOCIACAO = 5

# Colunas Membros
CARGO = 0
NOME_MEMBRO = 1
REPRESENTACAO = 2
RF_EOL = 3
CARGO_EDUCACAO = 4
EMAIL_MEMBRO = 5

# Linhas Membros
CARGOS = {
    'PRESIDENTE_DIRETORIA_EXECUTIVA': 1,
    'VICE_PRESIDENTE_DIRETORIA_EXECUTIVA': 2,
    'SECRETARIO': 3,
    'TESOUREIRO': 4,
    'VOGAL_1': 5,
    'VOGAL_2': 6,
    'VOGAL_3': 7,
    'VOGAL_4': 8,
    'VOGAL_5': 9,
    'PRESIDENTE_CONSELHO_FISCAL': 10,
    'CONSELHEIRO_1': 11,
    'CONSELHEIRO_2': 12,
    'CONSELHEIRO_3': 13,
    'CONSELHEIRO_4': 14,
}

# Colunas Contas da Associação
BANCO = 1
TIPO = 2
AGENCIA = 3
NUMERO = 4


def gerar_planilha(associacao):
    LOGGER.info(f'EXPORTANDO DADOS DA ASSOCIACAO {associacao.nome}...')

    path = os.path.join(os.path.basename(staticfiles_storage.location), 'modelos')
    nome_arquivo = os.path.join(path, 'modelo_exportacao_associacao.xlsx')
    workbook = load_workbook(nome_arquivo)

    flags = get_waffle_flag_model()

    dados_basicos(workbook, associacao)

    if flags.objects.filter(name='historico-de-membros', everyone=True).exists():
        membros_v2(workbook, associacao)
    else:
        membros(workbook, associacao)

    contas(workbook, associacao)

    return workbook


def dados_basicos(workbook, associacao):
    worksheet = workbook.worksheets[ASSOCIACAO]
    rows = list(worksheet.rows)
    rows[NOME_ASSOCIACAO][1].value = associacao.nome
    rows[EOL][1].value = associacao.unidade.codigo_eol
    rows[DRE][1].value = associacao.unidade.dre.nome if associacao.unidade.dre else ''
    rows[CNPJ][1].value = associacao.cnpj
    rows[CCM][1].value = associacao.ccm
    rows[EMAIL_ASSOCIACAO][1].value = associacao.email


def membros(workbook, associacao):
    membros = associacao.cargos.all()
    worksheet = workbook.worksheets[MEMBROS]
    rows = list(worksheet.rows)
    for membro in membros:
        linha = CARGOS[membro.cargo_associacao]
        rows[linha][NOME_MEMBRO].value = membro.nome
        rows[linha][REPRESENTACAO].value =  RepresentacaoCargo[membro.representacao].value
        rows[linha][RF_EOL].value = membro.codigo_identificacao
        rows[linha][CARGO_EDUCACAO].value = membro.cargo_educacao
        rows[linha][EMAIL_MEMBRO].value = membro.email


def membros_v2(workbook, associacao):
    from sme_ptrf_apps.mandatos.services import ServicoMandatoVigente
    from sme_ptrf_apps.mandatos.services import ServicoComposicaoVigente, ServicoCargosDaComposicao

    servico_mandato_vigente = ServicoMandatoVigente()
    mandato_vigente = servico_mandato_vigente.get_mandato_vigente()

    servico_composicao_vigente = ServicoComposicaoVigente(associacao=associacao, mandato=mandato_vigente)
    composicao_vigente = servico_composicao_vigente.get_composicao_vigente()

    servico_cargos_da_composicao = ServicoCargosDaComposicao(composicao=composicao_vigente)
    cargos_da_composicao = servico_cargos_da_composicao.get_cargos_da_composicao_ordenado_por_cargo_associacao()

    membros_da_composicao = []

    for cargo in cargos_da_composicao['diretoria_executiva']:
        membros_da_composicao.append({
            "cargo_associacao": cargo['cargo_associacao'],
            "nome": cargo["ocupante_do_cargo"]["nome"],
            "representacao": cargo["ocupante_do_cargo"]["representacao"],
            "codigo_identificacao": cargo["ocupante_do_cargo"]["codigo_identificacao"],
            "cargo_educacao": cargo["ocupante_do_cargo"]["cargo_educacao"],
            "email": cargo["ocupante_do_cargo"]["email"]
        })

    for cargo in cargos_da_composicao['conselho_fiscal']:
        membros_da_composicao.append({
            "cargo_associacao": cargo['cargo_associacao'],
            "nome": cargo["ocupante_do_cargo"]["nome"],
            "representacao": cargo["ocupante_do_cargo"]["representacao"],
            "codigo_identificacao": cargo["ocupante_do_cargo"]["codigo_identificacao"],
            "cargo_educacao": cargo["ocupante_do_cargo"]["cargo_educacao"],
            "email": cargo["ocupante_do_cargo"]["email"]
        })

    worksheet = workbook.worksheets[MEMBROS]
    rows = list(worksheet.rows)
    for membro_composicao in membros_da_composicao:
        linha = CARGOS[membro_composicao["cargo_associacao"]]
        rows[linha][NOME_MEMBRO].value = membro_composicao["nome"]
        rows[linha][REPRESENTACAO].value = RepresentacaoCargo[membro_composicao["representacao"]].value if membro_composicao["representacao"] else ''
        rows[linha][RF_EOL].value = membro_composicao["codigo_identificacao"]
        rows[linha][CARGO_EDUCACAO].value = membro_composicao["cargo_educacao"]
        rows[linha][EMAIL_MEMBRO].value = membro_composicao["email"]


def contas(workbook, associacao):
    contas = associacao.contas.all()
    worksheet = workbook.worksheets[CONTAS]
    linha = 2
    for conta in contas:
        worksheet.cell(row=linha, column=BANCO, value=conta.banco_nome)
        worksheet.cell(row=linha, column=TIPO, value=conta.tipo_conta.nome if conta.tipo_conta else ' ')
        worksheet.cell(row=linha, column=AGENCIA, value=conta.agencia)
        worksheet.cell(row=linha, column=NUMERO, value=conta.numero_conta)
        linha += 1


