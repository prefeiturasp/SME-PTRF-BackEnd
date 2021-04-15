import logging
import os
import re
from copy import copy
from datetime import date
from tempfile import NamedTemporaryFile

from django.contrib.staticfiles.storage import staticfiles_storage
from django.core.files import File
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter, range_boundaries

from sme_ptrf_apps.core.choices import MembroEnum
from sme_ptrf_apps.core.models import MembroAssociacao, RelacaoBens
from sme_ptrf_apps.despesas.models import RateioDespesa
from sme_ptrf_apps.despesas.tipos_aplicacao_recurso import APLICACAO_CAPITAL

from ..services.relacao_bens_dados_service import gerar_dados_relacao_de_bens
from ..services.relacao_bens_pdf_service import gerar_arquivo_relacao_de_bens_pdf

LOGGER = logging.getLogger(__name__)

COL_CABECALHO = 6
LINHA_PERIODO_CABECALHO = 4
LINHA_CONTA_CABECALHO = 5

# Colunas do blocos 2 da planilha
TIPO_DOCUMENTO = 0
NUMERO_DOCUMENTO = 1
DATA = 2
ESPECIFICACAO_MATERIAL = 3
NUMERO_DOCUMENTO_INCORPORACAO = 4
QUANTIDADE = 5
VALOR_ITEM = 6
VALOR_RATEIO = 7

BLOCO_3 = 19
LAST_LINE = 26


def gerar_arquivo_relacao_de_bens(periodo, conta_associacao, prestacao=None, previa=False):

    relacao_bens, _ = RelacaoBens.objects.update_or_create(
        conta_associacao=conta_associacao,
        prestacao_conta=prestacao,
        periodo_previa=None if prestacao else periodo,
        versao=RelacaoBens.VERSAO_PREVIA if previa else RelacaoBens.VERSAO_FINAL,
        status=RelacaoBens.STATUS_EM_PROCESSAMENTO,
    )

    filename = 'relacao_bens.xlsx'

    rateios = RateioDespesa.rateios_da_conta_associacao_no_periodo(
                conta_associacao=conta_associacao, periodo=periodo, aplicacao_recurso=APLICACAO_CAPITAL)

    if rateios:
        # PDF
        dados_relacao_de_bens = gerar_dados_relacao_de_bens(conta_associacao=conta_associacao, periodo=periodo, rateios=rateios)

        gerar_arquivo_relacao_de_bens_pdf(dados_relacao_de_bens=dados_relacao_de_bens, relacao_bens=relacao_bens)

        xlsx = gerar(periodo, conta_associacao, previa=previa)
        with NamedTemporaryFile() as tmp:
            xlsx.save(tmp.name)

            relacao_bens.arquivo.save(name=filename, content=File(tmp))
            relacao_bens.arquivo_concluido()

        return relacao_bens

    LOGGER.info("Não houve bem adquirido ou produzido no referido período (%s).", str(periodo))
    return None


def gerar(periodo, conta_associacao, previa=False):

    LOGGER.info("GERANDO RELAÇÃO DE BENS...")
    rateios = RateioDespesa.rateios_da_conta_associacao_no_periodo(
        conta_associacao=conta_associacao, periodo=periodo, aplicacao_recurso=APLICACAO_CAPITAL)
    path = os.path.join(os.path.basename(staticfiles_storage.location), 'cargas')
    nome_arquivo = os.path.join(path, 'modelo_relacao_de_bens.xlsx')
    workbook = load_workbook(nome_arquivo)
    worksheet = workbook.active

    cabecalho(worksheet, periodo, conta_associacao)
    identificacao_apm(worksheet, conta_associacao)
    data_geracao_documento(worksheet, previa)
    pagamentos(worksheet, rateios)

    return workbook


def cabecalho(worksheet, periodo, conta_associacao):
    rows = list(worksheet.rows)
    rows[LINHA_PERIODO_CABECALHO][COL_CABECALHO].value = str(periodo)
    rows[LINHA_CONTA_CABECALHO][COL_CABECALHO].value = conta_associacao.tipo_conta.nome


def identificacao_apm(worksheet, conta_associacao):
    """BLOCO 1 - IDENTIFICAÇÃO DA APM/APMSUAC DA UNIDADE EDUCACIONAL"""
    associacao = conta_associacao.associacao
    rows = list(worksheet.rows)
    rows[9][0].value = associacao.nome
    rows[9][4].value = associacao.cnpj
    rows[9][5].value = associacao.unidade.codigo_eol
    rows[9][6].value = associacao.unidade.dre.nome
    presidente_diretoria_executiva = MembroAssociacao.objects.filter(associacao=associacao,
                                                                     cargo_associacao=MembroEnum.PRESIDENTE_DIRETORIA_EXECUTIVA.name).first()
    rows[LAST_LINE-4][3].value = presidente_diretoria_executiva.nome if presidente_diretoria_executiva else ''

    autenticacao(worksheet, associacao)


def autenticacao(worksheet, associacao):
    """BLOCO 3 - AUTENTICAÇÃO"""

    mensagem = "Declaro, sob as penas de Lei, que os bens acima relacionados, adquiridos com recursos do PTRF foram doados" \
               " à Prefeitura do município de São Paulo/ Secretaria Municipal de Educação para serem incorporados ao patrimônio público" \
               f" e destinados à (ao) {associacao.unidade.tipo_unidade} {associacao.unidade.nome}, responsável por sua guarda e conservação."
    rows = list(worksheet.rows)
    rows[BLOCO_3][0].value = mensagem


def data_geracao_documento(worksheet, previa=False):
    rows = list(worksheet.rows)
    data_geracao = date.today().strftime("%d/%m/%Y")
    texto = f"Documento parcial gerado em: {data_geracao}" if previa else f"Documento final gerado em: {data_geracao}"
    rows[LAST_LINE - 1][0].value = texto


def pagamentos(worksheet, rateios, acc=0, start_line=15):
    """
    BLOCO 2 - IDENTIFICAÇÃO DOS BENS ADQUIRIDOS OU PRODUZIDOS
    """
    if not rateios:
        return

    quantidade = acc - 1 if acc else 0
    last_line = LAST_LINE + quantidade
    valor_total = sum(r.valor_rateio for r in rateios)
    row = list(worksheet.rows)[16]
    row[7].value = valor_total

    for linha, rateio in enumerate(rateios):
        # Movendo as linhas para baixo antes de inserir os dados novos
        ind = start_line + quantidade + linha
        if linha > 0:
            for row_idx in range(last_line + linha, ind - 2, -1):
                copy_row(worksheet, row_idx, 1, copy_data=True)

        row = list(worksheet.rows)[ind - 1]
        row[TIPO_DOCUMENTO].value = rateio.despesa.tipo_documento.nome if rateio.despesa.tipo_documento else ''
        row[NUMERO_DOCUMENTO].value = rateio.despesa.numero_documento
        row[DATA].value = rateio.despesa.data_documento.strftime("%d/%m/%Y") if rateio.despesa.data_documento else ''
        row[
            ESPECIFICACAO_MATERIAL].value = rateio.especificacao_material_servico.descricao if rateio.especificacao_material_servico else ''
        row[NUMERO_DOCUMENTO_INCORPORACAO].value = rateio.numero_processo_incorporacao_capital
        row[QUANTIDADE].value = rateio.quantidade_itens_capital
        row[VALOR_ITEM].value = rateio.valor_item_capital
        row[VALOR_RATEIO].value = rateio.valor_rateio


def copy_row(ws, source_row, dest_row, copy_data=False, copy_style=True, copy_merged_columns=True):
    """Copia uma linha da planilha para a linha imediatamento abaixo mantendo todos os atributos."""
    CELL_RE = re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    def replace(m):
        row = m.group('row')
        prefix = "$" if row.find("$") != -1 else ""
        row = int(row.replace("$", ""))
        row += dest_row if row > source_row else 0
        return m.group('col') + prefix + str(row)

    # Fazendo unmerge das celulas de destino.
    mergedcells = []
    for group in ws.merged_cells.ranges:
        mergedcells.append(group)

    for cr in mergedcells:
        min_col, min_row, max_col, max_row = range_boundaries(str(cr))
        if max_row == min_row == source_row + 1:
            ws.unmerge_cells(str(cr))

    new_row_idx = source_row + 1
    for row in range(new_row_idx, new_row_idx + dest_row):
        new_rd = copy(ws.row_dimensions[row - 1])
        new_rd.index = row
        ws.row_dimensions[row] = new_rd

        for col in range(ws.max_column):
            col = get_column_letter(col + 1)
            cell = ws['%s%d' % (col, row)]
            source = ws['%s%d' % (col, row - 1)]
            if copy_style:
                cell._style = copy(source._style)
            if copy_data and not isinstance(cell, MergedCell):
                cell.data_type = source.data_type
                cell.value = source.value

    for cr_idx, cr in enumerate(ws.merged_cell_ranges):
        ws.merged_cell_ranges[cr_idx] = CELL_RE.sub(
            replace,
            str(cr)
        )

    if copy_merged_columns:
        for cr in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cr))
            if max_row == min_row == source_row:
                for row in range(new_row_idx, new_row_idx + dest_row):
                    newCellRange = get_column_letter(min_col) + str(row) + ":" + get_column_letter(max_col) + str(row)
                    ws.merge_cells(newCellRange)


def apagar_previas_relacao_de_bens(periodo, conta_associacao):
    RelacaoBens.objects.filter(periodo_previa=periodo, conta_associacao=conta_associacao).delete()
