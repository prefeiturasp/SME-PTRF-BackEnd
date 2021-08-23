import logging
import os

from django.contrib.staticfiles.storage import staticfiles_storage
from openpyxl import load_workbook
import datetime

logger = logging.getLogger(__name__)

# Worksheets
SALDOS_BANCARIOS = 0

# Colunas
CODIGO_EOL = 1
NOME_UNIDADE = 2
TIPO_UNIDADE = 3
DATA_EXTRATO = 4
VALOR_EXTRATO = 5
PERIODO = 6
CONTA = 7
DRE = 8

FONTE = 1
GERADO_EM = 2
USUARIO = 3
FILTRO = 4



def gerar_planilha(dados):
    logger.info('Iniciando geração do arquivo xlsx...')

    path = os.path.join(os.path.basename(staticfiles_storage.location), 'modelos')
    nome_arquivo = os.path.join(path, 'modelo_exportacao_saldos_bancarios_associacoes.xlsx')
    workbook = load_workbook(nome_arquivo)

    extrair_dados(workbook, dados)

    return workbook


def extrair_dados(workbook, dados):

    worksheet = workbook.worksheets[SALDOS_BANCARIOS]
    linha = 2

    qs = dados["qs"]
    informacoes_adicionais = dados["informacoes_adicionais"]

    for dado in qs:
        worksheet.cell(row=linha, column=CODIGO_EOL, value=dado['unidade__codigo_eol'])
        worksheet.cell(row=linha, column=NOME_UNIDADE, value=dado['unidade__nome'])
        worksheet.cell(row=linha, column=TIPO_UNIDADE, value=dado['unidade__tipo_unidade'])
        worksheet.cell(row=linha, column=DATA_EXTRATO, value=formata_data(dado['obs_periodo__data_extrato']))
        worksheet.cell(row=linha, column=VALOR_EXTRATO, value=dado['obs_periodo__saldo_extrato'])
        worksheet.cell(row=linha, column=PERIODO, value=informacoes_adicionais['periodo'])
        worksheet.cell(row=linha, column=CONTA, value=informacoes_adicionais['conta'])
        worksheet.cell(row=linha, column=DRE, value=informacoes_adicionais['dre'])
        linha += 1

    linha += 3
    # Cria titulo das colunas adicionais
    worksheet.cell(row=linha, column=FONTE, value="Fonte")
    worksheet.cell(row=linha+1, column=FONTE, value="SIG-Escola/SME-SP")

    worksheet.cell(row=linha, column=GERADO_EM, value="Gerado em")
    worksheet.cell(row=linha+1, column=GERADO_EM, value=formata_data(datetime.datetime.today().strftime('%Y-%m-%d')))

    worksheet.cell(row=linha, column=USUARIO, value="Usuário")
    worksheet.cell(row=linha+1, column=USUARIO, value=informacoes_adicionais['usuario'])

    worksheet.cell(row=linha, column=FILTRO, value="Filtro")
    worksheet.cell(row=linha+1, column=FILTRO, value=informacoes_adicionais['filtro'])


def formata_data(data):
    if data is not None:
        original_date = datetime.datetime.strptime(str(data), '%Y-%m-%d')
        formatted_date = original_date.strftime("%d/%m/%Y")
        return formatted_date
    else:
        return None
