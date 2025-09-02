import csv
from datetime import datetime
import logging
from django.utils.timezone import make_aware

from django.core.files import File
from django.contrib.auth import get_user_model
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sme_ptrf_apps.core.models.ambiente import Ambiente
from sme_ptrf_apps.core.models.arquivos_download import ArquivoDownload
from sme_ptrf_apps.core.services.arquivo_download_service import (
    gerar_arquivo_download
)
from sme_ptrf_apps.utils.built_in_custom import get_recursive_attr

from tempfile import NamedTemporaryFile
from django.core.files.base import ContentFile
import os

logger = logging.getLogger(__name__)

CABECALHO_BENS = [
    ('Especificação dos bens', 'especificacao_bens'),
    ('Tipo de bem', 'tipo_bem'),
    ('Nº do processo de incorporação', 'numero_processo_incorporacao'),
    ('Tipo', 'tipo'),
    ('Número', 'numero'),
    ('Data', 'data'),
    ('Ação', 'acao'),
    ('Valor das despesas', 'valor_despesas'),
    ('Quantidade', 'quantidade'),
    ('Unitário', 'valor_unitario'),
    ('Valor (R$)', 'valor_total')
]


class ExportacaoConsultaBensProduzidosService:

    def __init__(self, **kwargs):
        logger.info("Inicializando ExportacaoConsultaBensProduzidosService")
        logger.info(f"Parâmetros recebidos: {kwargs}")
        
        self.queryset = kwargs.get('queryset', None)
        self.nome_arquivo = kwargs.get('nome_arquivo', None)
        self.data_inicio = kwargs.get("data_inicio", None)
        self.data_final = kwargs.get("data_final", None)
        self.cabecalho = CABECALHO_BENS
        self.ambiente = self.get_ambiente
        self.objeto_arquivo_download = None
        self.texto_filtro_aplicado = self.get_texto_filtro_aplicado()
        self.associacao = kwargs.get('associacao', None)
        self.codigo_eol = kwargs.get('codigo_eol', None)
        self.dre = kwargs.get('dre', None)
        self.filtros_str = kwargs.get('filtros_str', None)
        self.identificacao_usuario = kwargs.get('identificacao_usuario', None)
        
        # logger.info(f"Queryset: {self.queryset.count() if self.queryset else 'None'} registros")
        logger.info(f"Nome arquivo: {self.nome_arquivo}")
        logger.info(f"Data início: {self.data_inicio}")
        logger.info(f"Data final: {self.data_final}")
        logger.info(f"Usuário: {self.identificacao_usuario}")
        logger.info(f"Ambiente: {self.ambiente}")
        logger.info(f"Texto filtro: {self.texto_filtro_aplicado}")
        logger.info("ExportacaoConsultaBensProduzidosService inicializado com sucesso")

    @property
    def get_ambiente(self):
        ambiente = Ambiente.objects.first()
        prefixo = ambiente.prefixo if ambiente else ""
        logger.info(f"Ambiente obtido: {prefixo}")
        return prefixo

    def get_texto_filtro_aplicado(self):
        logger.info(f"Gerando texto do filtro - Data início: {self.data_inicio}, Data final: {self.data_final}")
        
        if self.data_inicio and self.data_final:
            data_inicio_formatada = datetime.strptime(f"{self.data_inicio}", '%Y-%m-%d')
            data_inicio_formatada = data_inicio_formatada.strftime("%d/%m/%Y")

            data_final_formatada = datetime.strptime(f"{self.data_final}", '%Y-%m-%d')
            data_final_formatada = data_final_formatada.strftime("%d/%m/%Y")

            texto = f"Filtro aplicado: {data_inicio_formatada} a {data_final_formatada} (por período)"
            logger.info(f"Texto filtro gerado: {texto}")
            return texto

        if self.data_inicio:
            data_inicio_formatada = datetime.strptime(f"{self.data_inicio}", '%Y-%m-%d')
            data_inicio_formatada = data_inicio_formatada.strftime("%d/%m/%Y")
            texto = f"Filtro aplicado: A partir de {data_inicio_formatada} (por período)"
            logger.info(f"Texto filtro gerado: {texto}")
            return texto

        if self.data_final:
            data_final_formatada = datetime.strptime(f"{self.data_final}", '%Y-%m-%d')
            data_final_formatada = data_final_formatada.strftime("%d/%m/%Y")
            texto = f"Filtro aplicado: Até {data_final_formatada} (por período)"
            logger.info(f"Texto filtro gerado: {texto}")
            return texto

        texto = ""
        logger.info(f"Texto filtro padrão: {texto}")
        return texto

    def exportar(self):
        """Método principal para exportação"""
        logger.info("Iniciando processo de exportação...")
        
        try:
            self.cria_registro_central_download()
            logger.info("Registro na central de download criado com sucesso")
            
            # Determinar o campo correto para filtro de data baseado no tipo de objeto
            if self.queryset and len(self.queryset) > 0:
                first_instance = self.queryset[0]
                logger.info(f"Primeira instância: {type(first_instance).__name__}")
                
                if hasattr(first_instance, 'criado_em'):
                    logger.info("Aplicando filtro por campo 'criado_em'")
                    self.filtra_range_data("criado_em")
                elif hasattr(first_instance, 'despesa') and hasattr(first_instance.despesa, 'data_documento'):
                    logger.info("Aplicando filtro por campo 'despesa__data_documento'")
                    self.filtra_range_data("despesa__data_documento")
                else:
                    # Se não houver campo de data, não aplicar filtro
                    logger.warning("Nenhum campo de data encontrado para filtro")
            else:
                logger.warning("Queryset vazio ou inexistente")
            
            logger.info("Iniciando exportação XLSX...")
            self.exportar_xlsx()
            logger.info("Exportação XLSX concluída com sucesso")
            
        except Exception as e:
            logger.error(f"Erro durante o processo de exportação: {e}")
            raise e

    def exportar_xlsx(self):
        """Exporta dados para formato XLSX"""
        try:
            logger.info("Iniciando exportação XLSX...")
            
            # Gerar a planilha base
            workbook = self.gerar_planilha_bens_produzidos()
            worksheet = workbook.active
            logger.info("Planilha base gerada com sucesso")
            
            # Atualiza o nome da associação
            cabecalho_associacao = worksheet['A5']
            cabecalho_associacao.value = self.associacao.nome
            
            # Atualiza o CNPJ
            cabecalho_cnpj = worksheet['D5']
            cabecalho_cnpj.value = self.associacao.cnpj

            # Atualiza o Código EOL
            cabecalho_eol = worksheet['G5']
            cabecalho_eol.value = self.codigo_eol
            
            # Atualiza o Diretoria Regional de Educação
            cabecalho_dre = worksheet['I5']
            cabecalho_dre.value = self.dre
            
            # Obter dados para exportação
            dados = self.monta_dados()
            logger.info(f"Dados montados: {len(dados)} linhas")
            
            if not dados:
                logger.warning("Nenhum dado para exportar - planilha será gerada apenas com cabeçalhos")
            
            # Adicionar dados na planilha a partir da linha 12
            linha_atual = 12
            for i, linha_dados in enumerate(dados):
                logger.info(f"Processando linha {i+1}: {linha_dados}")
                tipo_do_bem = linha_dados[1]
                
                # Alternar cor de fundo entre cinza e branco
                fill_cinza = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
                fill_branco = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                fill = fill_cinza if (i + 12) % 2 == 0 else fill_branco

                if tipo_do_bem == 'Produzido':
                    # Para bens produzidos, usar os dados já processados e colocar listas na vertical
                    # Extrair as listas de tipos, números, datas e ações
                    tipos = linha_dados[3] if len(linha_dados) > 3 else []
                    numeros = linha_dados[4] if len(linha_dados) > 4 else []
                    datas = linha_dados[5] if len(linha_dados) > 5 else []
                    acoes = linha_dados[6] if len(linha_dados) > 6 else []
                    valor_despesas = linha_dados[7] if len(linha_dados) > 7 else []
                    
                    # Determinar quantas linhas serão necessárias (máximo entre as listas)
                    max_linhas = max(len(tipos), len(numeros), len(datas), len(acoes))
                    if max_linhas == 0:
                        max_linhas = 1
                    
                    # Para cada item das listas, criar uma linha
                    for j in range(max_linhas):
                        # A: Especificação dos bens (coluna 1) - será mesclada se houver múltiplas linhas
                        cell_a = worksheet.cell(row=linha_atual, column=1)
                        cell_a.value = linha_dados[0] if len(linha_dados) > 0 else ''
                        cell_a.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_a.fill = fill
                        cell_a.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # B: Tipo de bem (coluna 2) - será mesclada se houver múltiplas linhas
                        cell_b = worksheet.cell(row=linha_atual, column=2)
                        cell_b.value = linha_dados[1] if len(linha_dados) > 1 else ''
                        cell_b.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_b.fill = fill
                        cell_b.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # C: Nº do processo de incorporação (coluna 3) - será mesclada se houver múltiplas linhas
                        cell_c = worksheet.cell(row=linha_atual, column=3)
                        cell_c.value = linha_dados[2] if len(linha_dados) > 2 else ''
                        cell_c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_c.fill = fill
                        cell_c.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # D: Tipo (coluna 4) - uma linha para cada tipo
                        cell_d = worksheet.cell(row=linha_atual, column=4)
                        cell_d.value = tipos[j] if j < len(tipos) else ''
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_d.fill = fill
                        cell_d.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # E: Número (coluna 5) - uma linha para cada número
                        cell_e = worksheet.cell(row=linha_atual, column=5)
                        cell_e.value = numeros[j] if j < len(numeros) else ''
                        cell_e.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_e.fill = fill
                        cell_e.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # F: Data (coluna 6) - uma linha para cada data
                        cell_f = worksheet.cell(row=linha_atual, column=6)
                        cell_f.value = datas[j] if j < len(datas) else ''
                        cell_f.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_f.fill = fill
                        cell_f.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # GH: Ação (colunas 7 e 8) - uma linha para cada ação
                        cell_g = worksheet.cell(row=linha_atual, column=7)
                        cell_g.value = acoes[j] if j < len(acoes) else ''
                        cell_g.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_g.fill = fill
                        cell_g.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        # Merge das colunas G e H
                        worksheet.merge_cells(start_row=linha_atual, start_column=7, end_row=linha_atual, end_column=8)
                        
                        # I: Valor das despesas (coluna 9) - será mesclada se houver múltiplas linhas
                        cell_i = worksheet.cell(row=linha_atual, column=9)
                        cell_i.value = valor_despesas[j] if j < len(valor_despesas) else '-'
                        cell_i.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_i.fill = fill
                        cell_i.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # J: Quantidade (coluna 10) - será mesclada se houver múltiplas linhas
                        cell_j = worksheet.cell(row=linha_atual, column=10)
                        cell_j.value = linha_dados[8] if len(linha_dados) > 8 else ''
                        cell_j.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_j.fill = fill
                        cell_j.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # K: Unitário (coluna 11) - será mesclada se houver múltiplas linhas
                        cell_k = worksheet.cell(row=linha_atual, column=11)
                        cell_k.value = linha_dados[9] if len(linha_dados) > 9 else ''
                        cell_k.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_k.fill = fill
                        cell_k.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # L: Valor (R$) (coluna 12) - será mesclada se houver múltiplas linhas
                        cell_l = worksheet.cell(row=linha_atual, column=12)
                        cell_l.value = linha_dados[10] if len(linha_dados) > 10 else ''
                        cell_l.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_l.fill = fill
                        cell_l.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        linha_atual += 1
                    
                    # Fazer merge das células para as colunas que devem ser mescladas quando há múltiplas linhas
                    if max_linhas > 1:
                        # Merge das colunas A, B, C, I, J, K, L (especificação, tipo de bem, processo, valor despesas, quantidade, unitário, valor total)
                        for col in [1, 2, 3, 10, 11, 12]:
                            start_row = linha_atual - max_linhas
                            end_row = linha_atual - 1
                            if start_row < end_row:
                                worksheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                                # Centralizar o conteúdo da célula mesclada
                                merged_cell = worksheet.cell(row=start_row, column=col)
                                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                else:
                    # Para bens adquiridos, manter o comportamento atual
                    # Extrair as listas de tipos, números e datas
                    tipos = linha_dados[3].split('; ') if linha_dados[3] else ['']
                    numeros = linha_dados[4].split('; ') if linha_dados[4] else ['']
                    datas = linha_dados[5].split('; ') if linha_dados[5] else ['']
                    
                    # Determinar quantas linhas serão necessárias (máximo entre as listas)
                    max_linhas = max(len(tipos), len(numeros), len(datas))
                    if max_linhas == 0:
                        max_linhas = 1
                    
                    # Para cada despesa, criar uma linha
                    for j in range(max_linhas):
                        # Mapear dados para as colunas corretas conforme estrutura definida
                        # A: Especificação dos bens (coluna 1) - será mesclada se houver múltiplas linhas
                        cell_a = worksheet.cell(row=linha_atual, column=1)
                        cell_a.value = linha_dados[0] if len(linha_dados) > 0 else ''
                        cell_a.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_a.fill = fill
                        cell_a.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # B: Tipo de bem (coluna 2) - será mesclada se houver múltiplas linhas
                        cell_b = worksheet.cell(row=linha_atual, column=2)
                        cell_b.value = linha_dados[1] if len(linha_dados) > 1 else ''
                        cell_b.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_b.fill = fill
                        cell_b.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # C: Nº do processo de incorporação (coluna 3) - será mesclada se houver múltiplas linhas
                        cell_c = worksheet.cell(row=linha_atual, column=3)
                        cell_c.value = linha_dados[2] if len(linha_dados) > 2 else ''
                        cell_c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_c.fill = fill
                        cell_c.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # D: Tipo (coluna 4) - uma linha para cada tipo
                        cell_d = worksheet.cell(row=linha_atual, column=4)
                        cell_d.value = tipos[j] if j < len(tipos) else ''
                        cell_d.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_d.fill = fill
                        cell_d.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # E: Número (coluna 5) - uma linha para cada número
                        cell_e = worksheet.cell(row=linha_atual, column=5)
                        cell_e.value = numeros[j] if j < len(numeros) else ''
                        cell_e.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_e.fill = fill
                        cell_e.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # F: Data (coluna 6) - uma linha para cada data
                        cell_f = worksheet.cell(row=linha_atual, column=6)
                        cell_f.value = datas[j] if j < len(datas) else ''
                        cell_f.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_f.fill = fill
                        cell_f.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # GH: Ação (colunas 7 e 8) - será mesclada se houver múltiplas linhas
                        cell_g = worksheet.cell(row=linha_atual, column=7)
                        cell_g.value = linha_dados[6] if len(linha_dados) > 6 else ''
                        cell_g.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_g.fill = fill
                        cell_g.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # Merge das colunas G e H
                        worksheet.merge_cells(start_row=linha_atual, start_column=7, end_row=linha_atual, end_column=8)
                        
                        # I: Valor das despesas (coluna 9) - será mesclada se houver múltiplas linhas
                        cell_i = worksheet.cell(row=linha_atual, column=9)
                        cell_i.value = linha_dados[7] if len(linha_dados) > 7 else ''
                        cell_i.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_i.fill = fill
                        cell_i.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # J: Quantidade (coluna 10) - será mesclada se houver múltiplas linhas
                        cell_j = worksheet.cell(row=linha_atual, column=10)
                        cell_j.value = linha_dados[8] if len(linha_dados) > 8 else ''
                        cell_j.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_j.fill = fill
                        cell_j.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # K: Unitário (coluna 11) - será mesclada se houver múltiplas linhas
                        cell_k = worksheet.cell(row=linha_atual, column=11)
                        cell_k.value = linha_dados[9] if len(linha_dados) > 9 else ''
                        cell_k.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_k.fill = fill
                        cell_k.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        # L: Valor (R$) (coluna 12) - será mesclada se houver múltiplas linhas
                        cell_l = worksheet.cell(row=linha_atual, column=12)
                        cell_l.value = linha_dados[10] if len(linha_dados) > 10 else ''
                        cell_l.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_l.fill = fill
                        cell_l.border = Border(
                            left=Side(style='thin', color='BFBFBF'),
                            right=Side(style='thin', color='BFBFBF'),
                            top=Side(style='thin', color='BFBFBF'),
                            bottom=Side(style='thin', color='BFBFBF')
                        )
                        
                        linha_atual += 1
                    
                    # Fazer merge das células para as colunas que devem ser mescladas quando há múltiplas linhas
                    if max_linhas > 1:
                        # Merge das colunas A, B, C, G, I, J, K, L (especificação, tipo de bem, processo, ação, valor despesas, quantidade, unitário, valor total)
                        for col in [1, 2, 3, 7, 9, 10, 11, 12]:
                            start_row = linha_atual - max_linhas
                            end_row = linha_atual - 1
                            if start_row < end_row:
                                worksheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
                                # Centralizar o conteúdo da célula mesclada
                                merged_cell = worksheet.cell(row=start_row, column=col)
                                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Adicionar texto de rodapé 4 linhas abaixo da última linha de dados
            rodape_linha = linha_atual + 4
            worksheet.merge_cells(f'A{rodape_linha}:L{rodape_linha}')
            rodape_celula = worksheet[f'A{rodape_linha}']
            # Formatar data e hora atual
            data_hora_atual = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            
            # Obter nome da associação
            nome_associacao = self.associacao.nome if self.associacao and self.associacao.nome else "N/A"
            
            # Obter cpf/rf do usuário
            username = self.identificacao_usuario if self.identificacao_usuario else "N/A"
            
            # Montar texto do rodapé no formato solicitado
            rodape_celula.value = f"{nome_associacao} - Documento final gerado pelo usuário {username}, via SIG - Escola, em: {data_hora_atual}"
            rodape_celula.font = Font(name='Calibri', color='000000', bold=False, size=11)
            rodape_celula.alignment = Alignment(horizontal='left', vertical='center')
            
            # OBS: Removido temporariamente para que possa ser usado futuramente, quando houver necessidade de exibir os filtros aplicados.
            # Adicionar texto dos filtros na linha 10, colunas B até L (Removido temporariamente)
            # filtros_celula = worksheet['B10']
            # filtros_celula.value = self.filtros_str if self.filtros_str else "Nenhum filtro aplicado"
            # filtros_celula.font = Font(name='Calibri', color='000000', bold=False, size=11)
            # filtros_celula.alignment = Alignment(horizontal='left', vertical='center')
            
            # Ajustar altura das linhas de dados
            for row in range(14, linha_atual):
                worksheet.row_dimensions[row].height = 20
            
            # Fazer merge das colunas GH em todas as linhas de dados (a partir da linha 14)
            if linha_atual > 14:  # Só fazer merge se houver dados
                for row in range(14, linha_atual):
                    worksheet.merge_cells(f'G{row}:H{row}')
                logger.info(f"Merge das colunas GH realizado em {linha_atual - 14} linhas de dados")
            
            logger.info(f"Dados inseridos na planilha até a linha {linha_atual-1}")
            
            # Salvar arquivo temporário com os parâmetros corretos
            logger.info("Salvando arquivo temporário...")
            with NamedTemporaryFile(
                suffix='.xlsx',
                delete=False
            ) as tmp:
                workbook.save(tmp.name)
                logger.info(f"Arquivo temporário salvo em: {tmp.name}")
                self.envia_arquivo_central_download(tmp)
                
        except Exception as e:
            logger.error(f"Erro ao exportar XLSX: {e}")
            if self.objeto_arquivo_download:
                self.objeto_arquivo_download.status = ArquivoDownload.STATUS_ERRO
                self.objeto_arquivo_download.msg_erro = str(e)
                self.objeto_arquivo_download.save()
            raise e

    def gerar_planilha_bens_produzidos(self):
        """Gera a planilha XLSX com formatação completa"""
        logger.info("Iniciando geração da planilha base...")
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Inventário de Bens"
        logger.info("Workbook criado com sucesso")

        # Inserir logo da prefeitura na linha 1
        try:
            logo_path = 'sme_ptrf_apps/static/images/logo-prefeitura-vertical.png'
            if os.path.exists(logo_path):
                img = Image(logo_path)
                img.width = 163
                img.height = 152
                worksheet.add_image(img, 'A1')
                logger.info('Logo inserida com sucesso')
            else:
                logger.warning(f'Logo não encontrada em: {logo_path}')
        except Exception as e:
            logger.error(f'Erro ao inserir logo: {e}')
            
        # Cor de fundo da A1 deve ser F6F8F9
        worksheet.cell(row=1, column=1).fill = PatternFill(start_color='F6F8F9', end_color='F6F8F9', fill_type='solid')
        # Ajustar altura da linha 1 para acomodar a logo
        worksheet.column_dimensions['A'].width = 144 / 7.0
        worksheet.row_dimensions[1].height = 110

        # Criar duas linhas separadas para o título com a mesma cor de fundo
        # Primeira linha: B1:L1
        worksheet.merge_cells('B1:J1')
        titulo_linha1 = worksheet['B1']
        titulo_linha1.value = "Programa de Transferência de Recursos Financeiros - PTRF\nInventário de Bens Adquiridos ou Produzidos"
        titulo_linha1.font = Font(name='Calibri', color='000000', bold=True, size=14)
        titulo_linha1.fill = PatternFill(start_color='F6F8F9', end_color='F6F8F9', fill_type='solid')
        titulo_linha1.alignment = Alignment(horizontal='left', vertical='center')
        
        # Texto descritivo: K1:L1
        worksheet.merge_cells('K1:L1')
        titulo_linha1 = worksheet['K1']
        titulo_linha1.value = "Inventário de Bens Adquiridos"
        titulo_linha1.font = Font(name='Calibri', color='000000', bold=True, size=12)
        titulo_linha1.fill = PatternFill(start_color='F6F8F9', end_color='F6F8F9', fill_type='solid')
        titulo_linha1.alignment = Alignment(horizontal='left', vertical='center')

        # Adicionar "Bloco 1 - Identificação da Associação" na linha 3
        worksheet.merge_cells('A3:L3')
        bloco1_cell = worksheet['A3']
        bloco1_cell.value = "Bloco 1 - Identificação da Associação"
        bloco1_cell.font = Font(name='Calibri', color='000000', bold=True, size=11)
        bloco1_cell.fill = PatternFill(start_color='F6F8F9', end_color='F6F8F9', fill_type='solid')
        bloco1_cell.alignment = Alignment(horizontal='left', vertical='center')

        # Aplicar bordas cinzas
        borda_cinza = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        bloco1_cell.border = borda_cinza

        # Garantir que a borda seja aplicada em toda a célula mesclada
        for col in range(1, 13):
            cell = worksheet.cell(row=3, column=col)
            cell.border = borda_cinza

        # Adicionar cabeçalhos e valores fictícios na linha 4 e 5
        # ABC: Nome da Associação
        worksheet.merge_cells('A4:C4')
        cabecalho_associacao = worksheet['A4']
        cabecalho_associacao.value = "Nome da Associação:"
        cabecalho_associacao.font = Font(name='Calibri', color='000000', bold=True, size=11)
        cabecalho_associacao.alignment = Alignment(horizontal='left', vertical='center')

        worksheet.merge_cells('A5:C5')
        valor_associacao = worksheet['A5']
        valor_associacao.value = "EMILIO RIBAS, PROFA."
        valor_associacao.font = Font(name='Calibri', color='000000', bold=False, size=11)
        valor_associacao.alignment = Alignment(horizontal='left', vertical='center')

        # DEF: CNPJ
        worksheet.merge_cells('D4:F4')
        cabecalho_cnpj = worksheet['D4']
        cabecalho_cnpj.value = "CNPJ:"
        cabecalho_cnpj.font = Font(name='Calibri', color='000000', bold=True, size=11)
        cabecalho_cnpj.alignment = Alignment(horizontal='left', vertical='center')

        worksheet.merge_cells('D5:F5')
        valor_cnpj = worksheet['D5']
        valor_cnpj.value = "12.123.123/0001-12"
        valor_cnpj.font = Font(name='Calibri', color='000000', bold=False, size=11)
        valor_cnpj.alignment = Alignment(horizontal='left', vertical='center')

        # GH: Código EOL
        worksheet.merge_cells('G4:H4')
        cabecalho_eol = worksheet['G4']
        cabecalho_eol.value = "Código EOL:"
        cabecalho_eol.font = Font(name='Calibri', color='000000', bold=True, size=11)
        cabecalho_eol.alignment = Alignment(horizontal='left', vertical='center')

        worksheet.merge_cells('G5:H5')
        valor_eol = worksheet['G5']
        valor_eol.value = "200222"
        valor_eol.font = Font(name='Calibri', color='000000', bold=False, size=11)
        valor_eol.alignment = Alignment(horizontal='left', vertical='center')

        # IJKL: Diretoria Regional de Educação
        worksheet.merge_cells('I4:L4')
        cabecalho_dre = worksheet['I4']
        cabecalho_dre.value = "Diretoria Regional de Educação:"
        cabecalho_dre.font = Font(name='Calibri', color='000000', bold=True, size=11)
        cabecalho_dre.alignment = Alignment(horizontal='left', vertical='center')

        worksheet.merge_cells('I5:L5')
        valor_dre = worksheet['I5']
        valor_dre.value = "GUAIANASES"
        valor_dre.font = Font(name='Calibri', color='000000', bold=False, size=11)
        valor_dre.alignment = Alignment(horizontal='left', vertical='center')

        # Aplicar bordas externas cinzas para a linha 4 (cabeçalhos)
        borda_externa_cinza = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )

        # Aplicar bordas externas para cada grupo de células mescladas da linha 4
        for col in range(1, 4):  # A4:C4
            cell = worksheet.cell(row=4, column=col)
            cell.border = borda_externa_cinza
        for col in range(4, 7):  # D4:F4
            cell = worksheet.cell(row=4, column=col)
            cell.border = borda_externa_cinza
        for col in range(7, 9):  # G4:H4
            cell = worksheet.cell(row=4, column=col)
            cell.border = borda_externa_cinza
        for col in range(9, 13):  # I4:L4
            cell = worksheet.cell(row=4, column=col)
            cell.border = borda_externa_cinza

        # Adicionar "Bloco 2 - Identificação dos Bens Adquiridos ou Produzidos" na linha 7
        worksheet.merge_cells('A7:L7')
        bloco2_cell = worksheet['A7']
        bloco2_cell.value = "Bloco 2 - Identificação dos Bens Adquiridos ou Produzidos"
        bloco2_cell.font = Font(name='Calibri', color='000000', bold=True, size=11)
        bloco2_cell.fill = PatternFill(start_color='F6F8F9', end_color='F6F8F9', fill_type='solid')
        bloco2_cell.alignment = Alignment(horizontal='left', vertical='center')
        bloco2_cell.border = borda_cinza

        # Garantir que a borda seja aplicada em toda a célula mesclada
        for col in range(1, 13):
            cell = worksheet.cell(row=7, column=col)
            cell.border = borda_cinza

        # OBS: Removido temporariamente para que possa ser usado futuramente, quando houver necessidade de exibir os filtros aplicados.
        # Adicionar "Filtros aplicados:" na linha 10, coluna A (Removido temporariamente)
        # filtros_cabecalho = worksheet['A10']
        # filtros_cabecalho.value = "Filtros aplicados:"
        # filtros_cabecalho.font = Font(name='Calibri', color='000000', bold=True, size=11)
        # filtros_cabecalho.alignment = Alignment(horizontal='left', vertical='center')

        # Adicionar texto dos filtros na linha 10, colunas B até L (Removido temporariamente)
        # worksheet.merge_cells('B10:L10')
        # filtros_valor = worksheet['B10']
        # filtros_valor.value = self.texto_filtro_aplicado
        # filtros_valor.font = Font(name='Calibri', color='000000', bold=False, size=11)
        # filtros_valor.alignment = Alignment(horizontal='left', vertical='center')

        # Adicionar cabeçalhos principais da tabela na linha 10
        self.criar_cabecalhos_principais(worksheet)

        # Adicionar cabeçalhos detalhados na linha 1
        self.criar_cabecalhos_detalhados(worksheet)

        # Ajustar larguras das colunas
        self.ajustar_larguras_colunas(worksheet)

        # Fixar as linhas de cabeçalho para rolagem
        worksheet.freeze_panes = 'A12'

        logger.info("Planilha base gerada com sucesso")
        return workbook

    def criar_cabecalhos_principais(self, worksheet):
        """Cria os cabeçalhos principais da tabela na linha 10"""
        logger.info("Criando cabeçalhos principais...")
        
        # Adicionar cabeçalho "Documento" na célula D10:F10
        worksheet.merge_cells('D10:F10')
        documento_cabecalho = worksheet['D10']
        documento_cabecalho.value = "Documento"
        documento_cabecalho.font = Font(name='Calibri', color='000000', bold=True, size=11)
        documento_cabecalho.alignment = Alignment(horizontal='center', vertical='center')

        # Adicionar cabeçalho "Detalhamento" na célula G10:I10
        worksheet.merge_cells('G10:I10')
        detalhamento_cabecalho = worksheet['G10']
        detalhamento_cabecalho.value = "Detalhamento"
        detalhamento_cabecalho.font = Font(name='Calibri', color='000000', bold=True, size=11)
        detalhamento_cabecalho.alignment = Alignment(horizontal='center', vertical='center')

        # Adicionar cabeçalho "Valor (R$)" na célula K10:L10
        worksheet.merge_cells('K10:L10')
        valor_cabecalho = worksheet['K10']
        valor_cabecalho.value = "Valor (R$)"
        valor_cabecalho.font = Font(name='Calibri', color='000000', bold=True, size=11)
        valor_cabecalho.alignment = Alignment(horizontal='center', vertical='center')

        # Aplicar bordas e cor de fundo para toda a linha 10
        self.aplicar_formato_linha(worksheet, 10)
        
        logger.info("Cabeçalhos principais criados com sucesso")

    def criar_cabecalhos_detalhados(self, worksheet):
        """Cria os cabeçalhos detalhados da tabela na linha 11"""
        # 1. Especificação dos bens (A11)
        worksheet['A11'].value = "1. Especificação dos bens"
        worksheet['A11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['A11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 2. Tipo de bem (B11)
        worksheet['B11'].value = "2. Tipo de bem"
        worksheet['B11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['B11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 3. Nº do processo de incorporação (C11)
        worksheet['C11'].value = "3. Nº do processo de incorporação"
        worksheet['C11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['C11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 4. Tipo (D11)
        worksheet['D11'].value = "4. Tipo"
        worksheet['D11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['D11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 5. Número (E11)
        worksheet['E11'].value = "5. Número"
        worksheet['E11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['E11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 6. Data (F11)
        worksheet['F11'].value = "6. Data"
        worksheet['F11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['F11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 7. Ação (G11:H11 - mesclado)
        worksheet.merge_cells('G11:H11')
        worksheet['G11'].value = "7. Ação"
        worksheet['G11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['G11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 8. Valor das despesas (I11)
        worksheet['I11'].value = "8. Valor das despesas"
        worksheet['I11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['I11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 9. Quantidade (J11)
        worksheet['J11'].value = "9. Quantidade"
        worksheet['J11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['J11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 10. Unitário (K11)
        worksheet['K11'].value = "10. Unitário"
        worksheet['K11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['K11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 11. Valor (R$) (L11)
        worksheet['L11'].value = "11. Valor (R$)"
        worksheet['L11'].font = Font(name='Calibri', color='000000', bold=True, size=11)
        worksheet['L11'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Aplicar bordas e cor de fundo para toda a linha 11
        self.aplicar_formato_linha(worksheet, 11)

    def aplicar_formato_linha(self, worksheet, linha):
        """Aplica formatação (bordas e cor de fundo) para uma linha específica"""
        borda_cabecalhos = Border(
            left=Side(style='thin', color='BFBFBF'),
            right=Side(style='thin', color='BFBFBF'),
            top=Side(style='thin', color='BFBFBF'),
            bottom=Side(style='thin', color='BFBFBF')
        )
        
        cor_fundo_cabecalhos = PatternFill(start_color='F6F8F9', end_color='F6F8F9', fill_type='solid')
        
        # Aplicar bordas e cor de fundo para toda a linha
        for col in range(1, 13):
            cell = worksheet.cell(row=linha, column=col)
            cell.border = borda_cabecalhos
            cell.fill = cor_fundo_cabecalhos

    def ajustar_larguras_colunas(self, worksheet):
        """Ajusta as larguras das colunas para acomodar melhor os cabeçalhos"""
        # worksheet.column_dimensions['A'].width = 18  # Especificação dos bens
        worksheet.column_dimensions['B'].width = 20  # Tipo de bem
        worksheet.column_dimensions['C'].width = 22  # Nº do processo de incorporação
        worksheet.column_dimensions['D'].width = 15  # Tipo
        worksheet.column_dimensions['E'].width = 15  # Número
        worksheet.column_dimensions['F'].width = 15  # Data
        worksheet.column_dimensions['G'].width = 20  # Ação (mesclado com H)
        worksheet.column_dimensions['H'].width = 20  # Ação (mesclado com G)
        worksheet.column_dimensions['I'].width = 18  # Valor das despesas
        worksheet.column_dimensions['J'].width = 15  # Quantidade
        worksheet.column_dimensions['K'].width = 15  # Unitário
        worksheet.column_dimensions['L'].width = 20  # Valor (R$)

    def monta_dados(self):
        """Monta os dados para exportação"""
        linhas_vertical = []
        
        if not self.queryset:
            logger.warning("Queryset vazio - nenhum dado para exportar")
            return linhas_vertical
        
        logger.info(f"Montando dados para {len(self.queryset)} registros")
        
        if len(self.queryset) == 0:
            logger.warning("Queryset vazio - nenhum dado para exportar")
            return linhas_vertical

        for instance in self.queryset:
            logger.info(f"Iniciando extração de dados, tipo: {type(instance).__name__}, id: {getattr(instance, 'id', 'N/A')}.")

            linha_horizontal = []
            
            # Determinar o tipo de objeto e mapear campos apropriadamente
            if hasattr(instance, 'despesa'):  # BemProduzidoItem
                logger.info(f"Processando bem adquirido: {instance}")
                linha_horizontal = self.monta_linha_bem_adquirido(instance)
            else:
                logger.info(f"Processando bem produzido: {instance}")
                linha_horizontal = self.monta_linha_bem_produzido(instance)

            if linha_horizontal:
                logger.info(f"Escrevendo linha {linha_horizontal}")
                linhas_vertical.append(linha_horizontal)
            else:
                logger.warning(f"Linha vazia para instância {instance}")

        print(linhas_vertical)
        logger.info(f"Total de linhas montadas: {len(linhas_vertical)}")
        return linhas_vertical

    def monta_linha_bem_produzido(self, instance):
        """Monta linha para bem produzido"""
        try:
            logger.info(f"Montando linha para bem produzido: {instance}")
            
            # 1. Especificação dos bens
            especificacao = getattr(instance.especificacao_do_bem, 'descricao', '') if instance.especificacao_do_bem else ''
            logger.info(f"Especificação: {especificacao}")
            
            # 2. Tipo de bem
            tipo_bem = 'Produzido'
            logger.info(f"Tipo de bem: Produzido")
            
            # 3. Nº do processo de incorporação
            processo = getattr(instance, 'num_processo_incorporacao', '-') or '-' if hasattr(instance, 'num_processo_incorporacao') else '-'
            logger.info(f"Processo: {processo}")
            
            # 4. Tipo (Documento) - Lista de tipos das despesas
            tipos = []

            if hasattr(instance, 'bem_produzido') and hasattr(instance.bem_produzido, 'despesas'):
                for bem_produzido_despesa in instance.bem_produzido.despesas.all():
                    tipo_doc = getattr(bem_produzido_despesa.despesa.tipo_documento, 'nome', '') if hasattr(bem_produzido_despesa.despesa, 'tipo_documento') else '-'
                    if tipo_doc:
                        tipos.append(tipo_doc)
            logger.info(f"Tipos: {tipos}")
            
            # 5. Número (Documento) - Lista de números das despesas
            numeros = []
            if hasattr(instance, 'bem_produzido') and hasattr(instance.bem_produzido, 'despesas'):
                for bem_produzido_despesa in instance.bem_produzido.despesas.all():
                    num_doc = getattr(bem_produzido_despesa.despesa, 'numero_documento', '') if hasattr(bem_produzido_despesa.despesa, 'numero_documento') else ''
                    if num_doc:
                        numeros.append(num_doc)
            logger.info(f"Números: {numeros}")
            
            # 6. Data (Data do Documento) - Lista de datas das despesas
            datas = []
            if hasattr(instance, 'bem_produzido') and hasattr(instance.bem_produzido, 'despesas'):
                for bem_produzido_despesa in instance.bem_produzido.despesas.all():
                    if hasattr(bem_produzido_despesa.despesa, 'data_documento') and bem_produzido_despesa.despesa.data_documento:
                        data_doc = bem_produzido_despesa.despesa.data_documento.strftime("%d/%m/%Y")
                        datas.append(data_doc)
            logger.info(f"Datas: {datas}")
            
            acoes = []
            if hasattr(instance, 'bem_produzido') and hasattr(instance.bem_produzido, 'despesas'):
                for bem_produzido_despesa in instance.bem_produzido.despesas.all():
                    if hasattr(bem_produzido_despesa.despesa, 'rateios'):
                        for rateio in bem_produzido_despesa.despesa.rateios.all():
                            if hasattr(rateio, 'acao_associacao') and rateio.acao_associacao:
                                if hasattr(rateio.acao_associacao, 'acao') and rateio.acao_associacao.acao:
                                    acoes.append(getattr(rateio.acao_associacao.acao, 'nome', '-'))
            logger.info(f"Ação: {acoes}")
                            
            # 8. Valor das despesas
            valor_despesas = []
            if hasattr(instance, 'bem_produzido') and hasattr(instance.bem_produzido, 'despesas'):
                for bem_produzido_despesa in instance.bem_produzido.despesas.all():
                    valor_despesas.append(bem_produzido_despesa.despesa.valor_total or '-')
            logger.info(f"Valor despesas: {valor_despesas}")
            
            # 9. Quantidade
            quantidade = getattr(instance, 'quantidade', 1) or 1
            logger.info(f"Quantidade: {quantidade}")
            
            # 10. Unitário - somar todos os valores das despesas primeiro
            valor_total_despesas = sum(valor_despesas) if valor_despesas else 0
            valor_unitario = valor_total_despesas / quantidade if quantidade > 0 else 0
            logger.info(f"Valor unitário: {valor_unitario}")
            
            # 11. Valor total
            valor_total = valor_total_despesas
            
            linha = [
                especificacao,
                tipo_bem,
                processo,
                tipos,
                numeros,
                datas,
                acoes,
                valor_despesas,
                str(quantidade),
                f"{valor_unitario:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"{valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ]
            
            logger.info(f"Linha montada com sucesso: {linha}")
            return linha
            
        except Exception as e:
            logger.error(f"Erro ao montar linha de bem produzido: {e}")
            return []

    def monta_linha_bem_adquirido(self, instance):
        """Monta linha para bem adquirido"""
        try:
            logger.info(f"Montando linha para bem adquirido: {instance}")
            
            # 1. Especificação dos bens
            especificacao = getattr(instance.especificacao_material_servico, 'descricao', '') if instance.especificacao_material_servico else ''
            logger.info(f"Especificação: {especificacao}")
            
            # 2. Tipo de bem
            tipo_bem = 'Adquirido'
            logger.info(f"Tipo de bem: Adquirido")
            
            # 3. Nº do processo de incorporação
            processo = getattr(instance, 'numero_processo_incorporacao_capital', '-') if hasattr(instance, 'numero_processo_incorporacao_capital') else '-'
            logger.info(f"Processo: {processo}")
            
            # 4. Tipo (Documento) - Lista de tipos das despesas
            tipos = []
            if hasattr(instance, 'despesa') and instance.despesa:
                # Para bens adquiridos, verificar se há uma despesa direta
                tipo_doc = getattr(instance.despesa.tipo_documento, 'nome', '') if hasattr(instance.despesa, 'tipo_documento') else ''
                if tipo_doc:
                    tipos.append(tipo_doc)
            # Verificar se há outras despesas relacionadas (se existir uma relação múltipla)
            if hasattr(instance, 'despesas') and hasattr(instance.despesas, 'all'):
                for despesa in instance.despesas.all():
                    tipo_doc = getattr(despesa.tipo_documento, 'nome', '') if hasattr(despesa, 'tipo_documento') else ''
                    if tipo_doc and tipo_doc not in tipos:
                        tipos.append(tipo_doc)
            tipo = '; '.join(tipos) if tipos else ''
            logger.info(f"Tipos: {tipo}")
            
            # 5. Número (Documento) - Lista de números das despesas
            numeros = []
            if hasattr(instance, 'despesa') and instance.despesa:
                # Para bens adquiridos, verificar se há uma despesa direta
                num_doc = getattr(instance.despesa, 'numero_documento', '') if hasattr(instance.despesa, 'numero_documento') else ''
                if num_doc:
                    numeros.append(num_doc)
            # Verificar se há outras despesas relacionadas (se existir uma relação múltipla)
            if hasattr(instance, 'despesas') and hasattr(instance.despesas, 'all'):
                for despesa in instance.despesas.all():
                    num_doc = getattr(despesa, 'numero_documento', '') if hasattr(despesa, 'numero_documento') else ''
                    if num_doc and num_doc not in numeros:
                        numeros.append(num_doc)
            numero = '; '.join(numeros) if numeros else ''
            logger.info(f"Números: {numero}")
            
            # 6. Data (Data do Documento) - Lista de datas das despesas
            datas = []
            if hasattr(instance, 'despesa') and instance.despesa:
                # Para bens adquiridos, verificar se há uma despesa direta
                if hasattr(instance.despesa, 'data_documento') and instance.despesa.data_documento:
                    data_doc = instance.despesa.data_documento.strftime("%d/%m/%Y")
                    datas.append(data_doc)
            # Verificar se há outras despesas relacionadas (se existir uma relação múltipla)
            if hasattr(instance, 'despesas') and hasattr(instance.despesas, 'all'):
                for despesa in instance.despesas.all():
                    if hasattr(despesa, 'data_documento') and despesa.data_documento:
                        data_doc = despesa.data_documento.strftime("%d/%m/%Y")
                        if data_doc not in datas:
                            datas.append(data_doc)
            data = '; '.join(datas) if datas else ''
            logger.info(f"Datas: {data}")
            
            # 7. Ação
            acao = ''
            if hasattr(instance, 'acao_associacao') and instance.acao_associacao:
                if hasattr(instance.acao_associacao, 'acao') and instance.acao_associacao.acao:
                    acao = getattr(instance.acao_associacao.acao, 'nome', '')
            logger.info(f"Ação: {acao}")
            
            # 8. Valor das despesas
            valor_despesas = instance.valor_rateio or 0
            logger.info(f"Valor despesas: {valor_despesas}")
            
            # 9. Quantidade
            quantidade = getattr(instance, 'quantidade', 1) or 1
            logger.info(f"Quantidade: {quantidade}")
            
            # 10. Unitário
            valor_unitario = valor_despesas / quantidade if quantidade > 0 else 0
            logger.info(f"Valor unitário: {valor_unitario}")
            
            # 11. Valor total
            valor_total = valor_despesas
            
            linha = [
                especificacao,
                tipo_bem,
                processo,
                tipo,
                numero,
                data,
                acao,
                f"{valor_despesas:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                str(quantidade),
                f"{valor_unitario:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'),
                f"{valor_total:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            ]
            
            logger.info(f"Linha montada com sucesso: {linha}")
            return linha
            
        except Exception as e:
            logger.error(f"Erro ao montar linha de bem adquirido: {e}")
            return []

    def filtra_range_data(self, field) -> list:
        """Filtra os dados por período de data"""
        import datetime
        
        logger.info(f"Aplicando filtro de data no campo: {field}")
        logger.info(f"Data início: {self.data_inicio}, Data final: {self.data_final}")
        logger.info(f"Queryset antes do filtro: {len(self.queryset)} registros")

        # Converte as datas inicial e final de texto para date
        inicio = datetime.datetime.strptime(self.data_inicio, "%Y-%m-%d").date() if self.data_inicio else None
        final = datetime.datetime.strptime(self.data_final, "%Y-%m-%d").date() if self.data_final else None

        # Define o horário da data_final para o último momento do dia
        final = make_aware(datetime.datetime.combine(final, datetime.time.max)) if final else None

        # Filtra a lista baseado no campo de data
        filtered_queryset = []
        
        for item in self.queryset:
            item_date = None
            
            # Extrair a data do item baseado no campo especificado
            if field == "criado_em" and hasattr(item, 'criado_em'):
                item_date = item.criado_em.date() if item.criado_em else None
            elif field == "despesa__data_documento" and hasattr(item, 'despesa') and hasattr(item.despesa, 'data_documento'):
                item_date = item.despesa.data_documento.date() if item.despesa.data_documento else None
            
            # Aplicar filtros de data
            if item_date:
                if inicio and final:
                    if inicio <= item_date <= final:
                        filtered_queryset.append(item)
                elif inicio and not final:
                    if item_date >= inicio:
                        filtered_queryset.append(item)
                elif final and not inicio:
                    if item_date <= final:
                        filtered_queryset.append(item)
                else:
                    filtered_queryset.append(item)
            else:
                # Se não tiver data, incluir no resultado
                filtered_queryset.append(item)

        self.queryset = filtered_queryset
        
        if inicio and final:
            logger.info(f"Filtro aplicado: entre {inicio} e {final}")
        elif inicio and not final:
            logger.info(f"Filtro aplicado: a partir de {inicio}")
        elif final and not inicio:
            logger.info(f"Filtro aplicado: até {final}")
        else:
            logger.info("Nenhum filtro de data aplicado")
        
        logger.info(f"Queryset após filtro: {len(self.queryset)} registros")
        return self.queryset

    def cria_registro_central_download(self):
        """Cria registro na central de download"""
        logger.info("Criando registro na central de download")
        logger.info(f"Usuário: {self.identificacao_usuario}")
        logger.info(f"Nome arquivo: {self.nome_arquivo}")
        logger.info(f"Texto filtro: {self.texto_filtro_aplicado}")

        usuario = get_user_model().objects.get(username=self.identificacao_usuario)

        obj = gerar_arquivo_download(
            usuario.username,
            self.nome_arquivo,
            self.texto_filtro_aplicado
        )

        self.objeto_arquivo_download = obj
        logger.info(f"Registro criado com ID: {obj.id}")

    def envia_arquivo_central_download(self, tmp):
        """Envia arquivo para a central de download"""
        try:
            logger.info("Salvando arquivo download...")
            logger.info(f"Arquivo temporário: {tmp.name}")
            logger.info(f"Tamanho do arquivo: {os.path.getsize(tmp.name)} bytes")
            
            # Para arquivos XLSX, precisamos ler o conteúdo do arquivo temporário
            with open(tmp.name, 'rb') as f:
                content = f.read()
            
            logger.info(f"Conteúdo lido: {len(content)} bytes")
            
            self.objeto_arquivo_download.arquivo.save(
                name=self.objeto_arquivo_download.identificador,
                content=File(ContentFile(content))
            )
            self.objeto_arquivo_download.status = ArquivoDownload.STATUS_CONCLUIDO
            self.objeto_arquivo_download.save()
            logger.info("Arquivo salvo com sucesso...")

        except Exception as e:
            logger.error(f"Erro arquivo download: {e}")
            self.objeto_arquivo_download.status = ArquivoDownload.STATUS_ERRO
            self.objeto_arquivo_download.msg_erro = str(e)
            self.objeto_arquivo_download.save()
            raise e

    def texto_rodape(self):
        """Gera texto do rodapé"""
        data_hora_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
        texto = f"Arquivo gerado via {self.ambiente} pelo usuário {self.identificacao_usuario} em {data_hora_geracao}"
        return texto

    def texto_info_arquivo_gerado(self):
        """Gera texto de informação do arquivo"""
        data_hora_geracao = datetime.now().strftime("%d/%m/%Y às %H:%M:%S")
        texto = f"Arquivo gerado via {self.ambiente} pelo usuário {self.identificacao_usuario} em {data_hora_geracao}"
        return texto


# Funções de compatibilidade para manter a interface existente
def gerar_planilha_bens_produzidos():
    """Função de compatibilidade - gera planilha básica"""
    service = ExportacaoConsultaBensProduzidosService()
    return service.gerar_planilha_bens_produzidos()


def salvar_planilha_bens_produzidos(caminho_arquivo):
    """Função de compatibilidade - salva planilha no caminho especificado"""
    try:
        workbook = gerar_planilha_bens_produzidos()
        workbook.save(caminho_arquivo)
        logger.info(f'Arquivo salvo com sucesso em: {caminho_arquivo}')
        return True
    except Exception as e:
        logger.error(f'Erro ao salvar arquivo: {e}')
        return False
